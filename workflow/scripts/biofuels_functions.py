
# Set working directory to be the project folder 
import os
import re
import pandas as pd 
import numpy as np
import glob
from datetime import datetime
from utility_functions import *
import yaml 
import shutil
from openpyxl import load_workbook
from openpyxl.comments import Comment

timestamp = datetime.now().strftime('%Y_%m_%d')
# wanted_wd = 'Outlook9th_EBT'
# os.chdir(re.split(wanted_wd, os.getcwd())[0] + wanted_wd)

def read_biofuels_input(biofuel_capacity_parameters_file_path):
    
    if not os.path.exists(biofuel_capacity_parameters_file_path):
        breakpoint()
        raise FileNotFoundError(f"Biorefining inputs Excel file not found: {biofuel_capacity_parameters_file_path}")
    
    workbook = pd.ExcelFile(biofuel_capacity_parameters_file_path)
    input_data_dict = {
        'capacity_df': pd.DataFrame(columns=['economy', 'fuel', 'year', 'additional_capacity_pj']),
        
        'utilisation_rate_df': pd.DataFrame(columns=['economy', 'fuel', 'year', 'utilisation_rate']),
        
        'EBT_biofuels_list': [],
        
        'simplified_economy_fuels': pd.DataFrame(columns=['economy', 'fuel', 'method']),
    }
    
    # Extract the EBT_biofuels_list and other input_data_dict data
    df = workbook.parse('config')
    input_data_dict['EBT_biofuels_list'] = df['EBT_biofuels_list'].dropna().tolist()
    
    # Read data from each sheet and reconstruct the YAML structure
    for sheet_name in workbook.sheet_names:
        df = workbook.parse(sheet_name)
        if '_capacity' in sheet_name:
            economy = sheet_name.replace('_capacity', '')
            df['economy'] = economy
            input_data_dict['capacity_df'] = pd.concat([input_data_dict['capacity_df'], df], ignore_index=True)
            
        elif sheet_name == 'config':
            continue
        elif sheet_name == 'simplified_economy_fuels':
            input_data_dict['simplified_economy_fuels'] = df
        elif sheet_name == 'utilisation_rate':
            economy = sheet_name.replace('_utilisation_rate', '')
            df['economy'] = economy
            input_data_dict['utilisation_rate_df'] = pd.concat([input_data_dict['utilisation_rate_df'], df], ignore_index=True)
                    
                    
    return input_data_dict
    
def create_biofuels_input_workbook(sheets_to_change = ['config', 'biofuels_capacity_additions', 'utilisation_rate', 'simplified_economy_fuels'], original_file_path = 'config/biofuel_capacity_parameters.xlsx', new_file_path = 'config/biofuel_capacity_parameters.xlsx', LOAD_AND_USE_ORIGINAL_FILE = True):
    """ONLY FOR PREPARATION OF THE BIOFUELS INPUT WORKBOOK. THIS WILL CREATE A NEW FILE WITH THE SAME SHEETS AS THE ORIGINAL BUT WITH THE DATA FILLED IN.
    RUN ME WITH 

        import biofuels_functions_new as biofuels_functions
        utils.set_working_directory()
        biofuels_functions.create_biofuels_input_workbook(sheets_to_change = ['config', 'biofuels_capacity_additions', 'utilisation_rate', 'simplified_economy_fuels'],original_file_path = 'config/biofuel_capacity_parameters_new.xlsx')
    """
    
    #first copy a version to archive/biofuel_capacity_parameters_{timestamp}.xlsx
    if not os.path.exists(original_file_path):
        raise FileNotFoundError(f'{original_file_path} not found')
    if not os.path.exists('archive'):
        os.makedirs('archive')
    timestamp = datetime.now().strftime('%Y_%m_%d')
    shutil.copy(original_file_path, f'archive/biofuel_capacity_parameters_{timestamp}.xlsx')
    # Open current workbook and edit the sheets
    if LOAD_AND_USE_ORIGINAL_FILE: 
        workbook = pd.ExcelFile(original_file_path, engine='openpyxl')
        workbook_sheet_names = workbook.sheet_names
    else:
        #create empty workbook object if we arent loading the original
        workbook_sheet_names = []
    writer = pd.ExcelWriter(new_file_path, engine='openpyxl')
    
    #add sheets that arent in sheets to change from the original file
    if LOAD_AND_USE_ORIGINAL_FILE: 
        for sheet_name in workbook_sheet_names:
            if sheet_name not in sheets_to_change:
                df = workbook.parse(sheet_name)
                df.to_excel(writer, sheet_name, index=False)
    #create a list of expected sheet names so that if any are missing we can add them:
    expected_sheets = ['config', 'utilisation_rate', 'simplified_economy_fuels'] + [f'{economy}_capacity' for economy in ALL_ECONOMY_IDS]
    completed_sheets = []
    for sheet_name in workbook_sheet_names + expected_sheets:
        if sheet_name in completed_sheets:
            continue
        completed_sheets += [sheet_name]
        if ('biofuels_capacity_additions' in sheets_to_change or sheet_name not in workbook_sheet_names) and '_capacity' in sheet_name:
            # create data for 01_AUS as an example. set all values to 0 with the columns EBT_fuel	additional_energy_pj	economy	scenario	specific_fuel	year
            #create a row for each scenario in SCENARIOS_list, a fuel for each one in 16_05_biogasoline 16_06_biodiesel 16_07_bio_jet_kerosene 16_01_biogas 15_01_fuelwood_and_woodwaste 15_02_bagasse 15_03_charcoal 15_04_black_liquor 15_05_other_biomass, name the spoeific fuel the same as the EBT_fuel without the number at the start, and set the year to OUTLOOK_BASE_YEAR
            df = pd.DataFrame(columns=['EBT_fuel', 'additional_capacity_pj', 'economy', 'scenario', 'specific_fuel', 'year'])
            economy = sheet_name.replace('_capacity', '')
            for scenario in SCENARIOS_list:
                for fuel in ['16_05_biogasoline', '16_06_biodiesel', '16_07_bio_jet_kerosene', '16_01_biogas', '15_01_fuelwood_and_woodwaste', '15_02_bagasse', '15_03_charcoal', '15_04_black_liquor', '15_05_other_biomass']:
                    #remove all numbers from fuel name then strip off the _'s at the start
                    specific_fuel = re.sub(r'\d', '', fuel).lstrip('_')
                    df = pd.concat([df, pd.DataFrame([{'EBT_fuel': fuel, 'additional_capacity_pj': 0, 'economy': economy, 'scenario': scenario, 'specific_fuel': specific_fuel, 'year': OUTLOOK_BASE_YEAR}])], ignore_index=True)
                    
            df.to_excel(writer, sheet_name, index=False)
        elif ('utilisation_rate' in sheets_to_change or sheet_name not in workbook_sheet_names) and sheet_name == 'utilisation_rate':
            # create data for 01_AUS as an example. set all values to 1 with the columns economy	utilisation_rate	scenario	year	fuel
            #create a row for each scenario in SCENARIOS_list, a fuel for each one in 16_05_biogasoline 16_06_biodiesel 16_07_bio_jet_kerosene 16_01_biogas 15_01_fuelwood_and_woodwaste 15_02_bagasse 15_03_charcoal 15_04_black_liquor 15_05_other_biomass, and set the year to OUTLOOK_BASE_YEAR
            df = pd.DataFrame(columns=['economy', 'utilisation_rate', 'scenario', 'year', 'fuel'])
            for scenario in SCENARIOS_list:
                for fuel in ['16_05_biogasoline', '16_06_biodiesel', '16_07_bio_jet_kerosene', '16_01_biogas', '15_01_fuelwood_and_woodwaste', '15_02_bagasse', '15_03_charcoal', '15_04_black_liquor', '15_05_other_biomass']:
                    df = pd.concat([df, pd.DataFrame([{'economy': '01_AUS', 'utilisation_rate': 1, 'scenario': scenario, 'year': OUTLOOK_BASE_YEAR, 'fuel': fuel}])], ignore_index=True)
                    
            df.to_excel(writer, sheet_name, index=False)
        elif ('config' in sheets_to_change or sheet_name not in workbook_sheet_names) and sheet_name == 'config':
            #create data for EBT_biofuels_list 
            df = pd.DataFrame(columns=['EBT_biofuels_list'])
            for fuel in ['16_05_biogasoline', '16_06_biodiesel', '16_07_bio_jet_kerosene', '16_01_biogas', '15_01_fuelwood_and_woodwaste', '15_02_bagasse', '15_03_charcoal', '15_04_black_liquor', '15_05_other_biomass']:
                df = pd.concat([df, pd.DataFrame([{'EBT_biofuels_list': fuel}])], ignore_index=True)
            df.to_excel(writer, sheet_name, index=False)
                
        elif ('simplified_economy_fuels' in sheets_to_change or sheet_name not in workbook_sheet_names) and sheet_name == 'simplified_economy_fuels': 
            # #create example but using 00_mars as the economy. and create a version for ewach of methods: satisfy_all_demand_with_domestic_production, satisfy_all_demand_with_domestic_production_RAMP, keep_same_ratio_of_production_to_consumption
            
            # df.to_excel(writer, sheet_name, index=False)
            df = pd.DataFrame(columns=['economy', 'fuel', 'method'])
            for fuel in ['16_05_biogasoline', '16_06_biodiesel', '16_07_bio_jet_kerosene', '16_01_biogas', 
                         '15_01_fuelwood_and_woodwaste', '15_02_bagasse', '15_03_charcoal', '15_04_black_liquor', 
                         '15_05_other_biomass']:
                for method in ['satisfy_all_demand_with_domestic_production', 
                               'satisfy_all_demand_with_domestic_production_RAMP', 
                               'keep_same_ratio_of_production_to_consumption']:
                    df = pd.concat([df, pd.DataFrame([{'economy': '00_MARS', 'fuel': fuel, 'method': method}])], ignore_index=True)
                    if method == 'satisfy_all_demand_with_domestic_production_RAMP':
                        df = pd.concat([df, pd.DataFrame([{'economy': '00_MARS', 'fuel': fuel, 'method': 'satisfy_all_demand_with_domestic_production_RAMP20'}])], ignore_index=True)
                #but then also, for every economy we want to have 'keep_same_ratio_of_production_to_consumption' for every fuel as the default method, so insert those
                for economy in ALL_ECONOMY_IDS:
                    df = pd.concat([df, pd.DataFrame([{'economy': economy, 'fuel': fuel, 'method': 'keep_same_ratio_of_production_to_consumption'}])], ignore_index=True)
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
        elif sheet_name in workbook_sheet_names:
            #save as is
            df = workbook.parse(sheet_name)
            df.to_excel(writer, sheet_name, index=False)
        else:
            breakpoint()
            raise Exception(f'No code to create data for {sheet_name}')
    writer.close()
    #if we updated simplified_economy_fuels then add a comment with the explanation of the methods
    if 'simplified_economy_fuels' in sheets_to_change:
        explanation = (
            "Explanation of Methods:\n\n"
            "- satisfy_all_demand_with_domestic_production:\n"
            "    The economy satisfies all demand for a fuel with domestic production. "
            "This eliminates imports and exports, making the economy self-sufficient.\n\n"
            "- satisfy_all_demand_with_domestic_production_RAMP:\n"
            "    Gradually increases domestic production over time to meet 100% of demand, "
            "avoiding sudden jumps in production.\n\n"
            "- keep_same_ratio_of_production_to_consumption:\n"
            "    Maintains the existing ratio of exports, imports, and production, while "
            "ensuring all demand is met with domestic production."
        )
        
        add_comment_to_sheet(new_file_path, 'simplified_economy_fuels', "C1", explanation)
    return None
            

def add_comment_to_sheet(file_path, sheet_name, cell, comment_text):
        
    from openpyxl import load_workbook
    from openpyxl.comments import Comment
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
    
    ws = wb[sheet_name]
    comment = Comment(comment_text, "Explanation")
    ws[cell].comment = comment
    wb.save(file_path)
    wb.close()
    return None

def biofuels_supply_and_transformation_handler(economy, model_df_clean_wide, PLOT = True, biofuel_capacity_parameters_file_path = 'config/biofuel_capacity_parameters.xlsx'):

    """to save workload we are just going to do biofuels production and all biofuels supply within this model. it will take in demand for biofuels and then depending on a few paramerters, it will output production, imports and exports of biofuels. 
    
    There is also the option to include expected production capacity. and also an option to set an economy to just continue with the same ratio of exports/imports/production and satisfy all demand with domestic production. 
     
    The parameters will be: 
    capacity additions (of each biofuel)
    
    
    This will also have the capcity to output the total amount of feedstock required and detailed information on the type of capacity (which will have to be detailed within the input params)
    
    It will also need to calcualte capacity based on current production. 
    
    The parameter format will be:
    
    ```yml
    biofuels:
        - economy: 01_AUS
            capacity_additions:
            - year: 2030
                additional_energy_pj: 0
                specific_fuel: 'ethanol'
                EBT_fuel: '16_05_biogasoline'
    ```
    Also there will be this in the yaml for biorefinery utilisation rates:
    ```yml
    utilisation_rate:
        - economy: 01_AUS
            initial_avg_utilisation_rate: 1
            utilisation_rate_changes:
            - year: 2030
                new_avg_utilisation_rate: 1
    ```                
    So yeah. 
    """
        
    input_data_dict = read_biofuels_input(biofuel_capacity_parameters_file_path)
    
    capacity_df, detailed_capacity_df, production_df = prepare_capacity_data(economy, model_df_clean_wide, input_data_dict)

    consumption_df = prepare_consumption_data(economy, model_df_clean_wide, input_data_dict)
<<<<<<< HEAD

=======
    
>>>>>>> d17335e41b94593dde28c4249295c1c398b475b6
    production_df, consumption_df = biofuel_simplified_modelling_methods_handler(production_df, consumption_df, input_data_dict, model_df_clean_wide)
    
    final_df = calculate_exports_imports(model_df_clean_wide, consumption_df, production_df)#NOT 100% URE THATIF THERE IS NO CAP DATA FOR AN ECONOMY THAT simplified_economy_fuels WILL BE USED?
    if PLOT and final_df.shape[0] > 0:
        print('Plotting biofuels data')
        plot_biofuels_data(final_df, economy)
        
    save_results(final_df, detailed_capacity_df, capacity_df, economy)
    ###########################################################
    ###########################################################

def prepare_capacity_data(economy, model_df_clean_wide, input_data_dict):   
    #find the economy in the input_data_dict
    economy_capacity_df = input_data_dict['capacity_df']
    economy_capacity_df = economy_capacity_df[economy_capacity_df.economy == economy]
    simplified_economy_fuels = input_data_dict['simplified_economy_fuels']
    if economy_capacity_df.shape[0] == 0 and economy not in simplified_economy_fuels['economy'].unique():
        raise Exception(f'No capacity data for {economy} in the biofuel_params and it is not in the simplified_economy_fuels either')
    
    #dso the same for the utiisation rates
    utilisation_rate = input_data_dict['utilisation_rate_df']
    utilisation_rate = utilisation_rate[utilisation_rate.economy == economy]#not important if empty since we can default to 1
    
    #create a dataframe to store the capacity values for each biofuel type
    EBT_biofuels_list = input_data_dict['EBT_biofuels_list']
    capacity_df = pd.DataFrame(columns = ['economy', 'year', 'scenarios']+EBT_biofuels_list)
    utilisations_df = pd.DataFrame(columns = ['economy', 'year', 'scenarios']+EBT_biofuels_list)
    
    #fill with all the years:
    latest_hist = OUTLOOK_BASE_YEAR
    proj_years = list(range(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR+1, 1))
    historical_years = list(range(EBT_EARLIEST_YEAR, OUTLOOK_BASE_YEAR+1, 1))
    # Convert year columns to integers if they are strs
    
    # model_df_clean_wide.columns = [int(col) if re.search(r'\d{4}', col) else col for col in model_df_clean_wide.columns]
    model_df_clean_wide.columns = [int(col) if re.search(r'\d{4}', str(col)) else col for col in model_df_clean_wide.columns]
    
    capacity_df['year'] = [latest_hist] + proj_years
    capacity_df['economy'] = economy
    capacity_df[EBT_biofuels_list] = 0
    utilisations_df['year'] = [latest_hist] + proj_years
    utilisations_df['economy'] = economy
    utilisations_df[EBT_biofuels_list] = np.nan
    
    #make equal sized capcity df for each scnerio in model_df_clean_wide:
    capacity_df_new = pd.DataFrame(columns = ['economy', 'year', 'scenarios']+EBT_biofuels_list)
    for scenario in model_df_clean_wide['scenarios'].unique():
        capacity_df['scenarios'] = scenario
        capacity_df_new = pd.concat([capacity_df_new, capacity_df])
    capacity_df = capacity_df_new.copy().reset_index(drop = True)
    
    utilisations_df_new = pd.DataFrame(columns = ['economy', 'year', 'scenarios']+EBT_biofuels_list)
    for scenario in model_df_clean_wide['scenarios'].unique():
        utilisations_df['scenarios'] = scenario
        utilisations_df_new = pd.concat([utilisations_df_new, utilisations_df])
    utilisations_df = utilisations_df_new.copy().reset_index(drop = True)
        
    #now for each capacity addition, add the capacity to the capacity_df:
    detailed_capacity_df = pd.DataFrame(columns = ['economy', 'year','scenarios', 'specific_fuel', 'EBT_fuel', 'capacity_type', 'capacity'])
    
    for i, row in economy_capacity_df.iterrows():
        year = row['year']
        fuel = row['EBT_fuel']
        scenario = row['scenario']
        if fuel not in EBT_biofuels_list:
            raise Exception(f'{fuel} not in EBT_biofuels_list. Need to add it to the yaml file')
        capacity_df.loc[(capacity_df['year'] == year) & (capacity_df['scenarios'] == scenario), fuel] += row['additional_capacity_pj']
        
        #add the detailed capacity to the detailed_capacity_df
        specific_fuel = row['specific_fuel']
        capacity_type = 'new_capacity'
        capacity = row['additional_capacity_pj']
        if capacity != 0:#this allows for negative capacity too! (and theres checks later to make sure production is never negative)
            detailed_capacity_df = pd.concat([detailed_capacity_df, pd.DataFrame([{'economy': economy, 'year': year, 'specific_fuel': specific_fuel, 'EBT_fuel': fuel, 'capacity_type': capacity_type, 'capacity': capacity, 'scenarios': scenario}])], ignore_index=True)
            
    for i, row in utilisation_rate.iterrows():
        #economy	utilisation_rate	scenario	year	fuel
        year = row['year']
        fuel = row['fuel']
        scenario = row['scenario']
        utilisation_rate = row['utilisation_rate']
        if fuel not in EBT_biofuels_list:
            raise Exception(f'{fuel} not in EBT_biofuels_list. Need to add it to the yaml file')
        utilisations_df.loc[(utilisations_df['year'] == year) & (utilisations_df['scenarios'] == scenario), fuel] = utilisation_rate
    #set base year in detailed_capacity_dfa capacity df to 1 if its not set already
    
    for fuel in EBT_biofuels_list:
        if pd.isna(utilisations_df.loc[(utilisations_df['economy'] == economy) & (utilisations_df['year'] == latest_hist), fuel]).all():
            utilisations_df.loc[(utilisations_df['economy'] == economy) & (utilisations_df['year'] == latest_hist), fuel] = 1
    utilisations_df = utilisations_df.ffill()
    
    #check for no duplicates
    if utilisations_df.duplicated(subset = ['economy', 'year', 'scenarios']).sum() > 0:
        raise Exception('There are duplicates in the utilisations_df')
    
    #calcaulte the capcity for the historical years. This will require calcaulting total production of each biofuel type and then working out the capacity required to produce that amount (convert it at a 100% utilisation rate so that 1 PJ of production requires 1 PJ of capacity)
    #grab the production data from the model_df_clean_wide
    # scenarios	economy	sectors	sub1sectors	sub2sectors	sub3sectors	sub4sectors	fuels	subfuels	subtotal_layout	subtotal_results
    
    biofuel_production_historical = model_df_clean_wide[(model_df_clean_wide['sectors'] == '01_production') & ((model_df_clean_wide['subfuels'].isin(EBT_biofuels_list) | model_df_clean_wide['fuels'].isin(EBT_biofuels_list)) & (model_df_clean_wide['subtotal_layout'] == False))].copy().reset_index(drop = True)
    #keep only the columns for historical years
    biofuel_production_historical = biofuel_production_historical[['scenarios','economy','fuels', 'subfuels']+historical_years]
    #sum
    biofuel_production_historical = biofuel_production_historical.groupby(['scenarios','economy','fuels','subfuels']).sum().reset_index()
    #where subfuels is 'x' set it to fuels
    biofuel_production_historical['subfuels'] = np.where(biofuel_production_historical['subfuels'] == 'x', biofuel_production_historical['fuels'], biofuel_production_historical['subfuels']) 
    #drop fuels
    biofuel_production_historical.drop(columns = 'fuels', inplace = True)
    #melt and then pivot so the years are in one column and fuels are the columns
    biofuel_production_historical = biofuel_production_historical.melt(id_vars = ['scenarios','economy','subfuels'], value_vars = historical_years, var_name = 'year', value_name = 'production')
    biofuel_production_historical = biofuel_production_historical.pivot(index = ['scenarios','economy','year'], columns = 'subfuels', values = 'production').reset_index()
    #double check all the values in EBT_biofuels_list are in the columns
    for fuel in EBT_biofuels_list:
        if fuel not in biofuel_production_historical.columns:
            raise Exception(f'{fuel} not in biofuel_production_historical columns')
        
    #now we need to calculate the capacity for each year by adding the cumulative total for each fuel type and thwn later dividing by the utilisation rate
    
    #concatenate the historical and future capacity dfs and recalcaulta the total capacity for each year
    #first check cols are the same
    if set(capacity_df.columns) != set(biofuel_production_historical.columns):
        breakpoint()
        raise Exception('The columns in capacity_df and biofuel_production_historical are not the same')
    
    #drop BASE_YEAR from capacity_df
    capacity_df = capacity_df[capacity_df['year'] != latest_hist].copy().reset_index(drop = True)
    
    #concatenate
    capacity_df = pd.concat([capacity_df, biofuel_production_historical]).copy().reset_index(drop = True)
    
    #now groupby and cumsum the dates incl and after the base year
    capacity_df.sort_values(by = ['year'], inplace = True)
    #set any nas to 0
    capacity_df[EBT_biofuels_list] = capacity_df[EBT_biofuels_list].fillna(0)
    capacity_df[EBT_biofuels_list] = capacity_df[EBT_biofuels_list].apply(pd.to_numeric, errors='coerce')
    
    capacity_df_projected = capacity_df[capacity_df['year'] >= latest_hist].copy().reset_index(drop = True)
    capacity_df_projected[EBT_biofuels_list] = capacity_df_projected.groupby(['scenarios', 'economy'])[EBT_biofuels_list].cumsum()
    capacity_df = pd.concat([capacity_df[capacity_df['year'] < latest_hist], capacity_df_projected]).copy().reset_index(drop = True)
    
    #check fr nas
    if capacity_df[EBT_biofuels_list].isnull().sum().sum() > 0:
        breakpoint()
        raise Exception('There are NaNs in the capacity_df')
    
    # now calcuate production by usign the capacity and utilisation rate dfs:
    production_df = capacity_df.merge(utilisations_df, on = ['economy', 'year', 'scenarios'], how = 'left', suffixes=('', '_ur')).copy().reset_index(drop = True)
    #since utilisations sf doesnt have values for pre-base year we need to fill them with 1    
    for fuel in EBT_biofuels_list:
        production_df[fuel] = production_df[fuel] * production_df[fuel+'_ur'].fillna(1)
    #drop the utilisation rate columns
    production_df.drop(columns = [fuel+'_ur' for fuel in EBT_biofuels_list], inplace = True)
    #melt so fuel is a column
    production_df = production_df.melt(id_vars = ['scenarios', 'economy','year'], value_vars = EBT_biofuels_list, var_name = 'subfuels', value_name = 'production')
    
    #set any prodcuton less than 0 to 0
    production_df['production'] = np.where(production_df['production'] < 0, 0, production_df['production'])
    
        
    return capacity_df, detailed_capacity_df, production_df

def prepare_consumption_data(economy, model_df_clean_wide, input_data_dict):
    EBT_biofuels_list = input_data_dict['EBT_biofuels_list']
    proj_years = list(range(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR+1, 1))
    historical_years = list(range(EBT_EARLIEST_YEAR, OUTLOOK_BASE_YEAR+1, 1))
    all_years = historical_years + proj_years
    all_years = [year for year in all_years if year in model_df_clean_wide.columns]
    #now we need to calculate imports and exports. This will be defined by the amount of consumption of each fuel in the demand and power sectors:
    #grab the consumption data from the model_df_clean_wide for future years
    sectors= ['04_international_marine_bunkers', '05_international_aviation_bunkers', '09_total_transformation_sector', '10_losses_and_own_use', '14_industry_sector', '15_transport_sector', '16_other_sector', '17_nonenergy_use']
    removed_column_value_dict = {'sub1sectors':'09_10_biofuels_processing'}
    consumption_df = model_df_clean_wide[(model_df_clean_wide['sectors'].isin(sectors)) & (model_df_clean_wide['subfuels'].isin(EBT_biofuels_list)) & (model_df_clean_wide['subtotal_results'] == False)].copy().reset_index(drop = True)
    print('Consumption data for biofuels is based on the sectors {}'.format(sectors))
    for column, value in removed_column_value_dict.items():
        consumption_df = consumption_df[consumption_df[column] != value].copy().reset_index(drop = True)
        print('Removing {} from {} in the biofuels consumption calculation for biofuels'.format(value, column))
    
    #drop any postivie values in 09_total_transformation_sector and then set any negative values to positive
    for year in all_years:
        consumption_df.loc[(consumption_df['sectors'] == '09_total_transformation_sector') & (consumption_df[year] > 0), year] = 0
        consumption_df[year] = consumption_df[year].abs()
    
    consumption_df = consumption_df[['scenarios','economy','subfuels']+all_years]
    #sum
    consumption_df = consumption_df.groupby(['scenarios','economy','subfuels']).sum().reset_index()
    #melt and then pivot so the years are in one column and fuels are the columns. this makes it compatible with the production data so we can calculate imports and exports
    consumption_df = consumption_df.melt(id_vars = ['scenarios','economy','subfuels'], value_vars =all_years, var_name = 'year', value_name = 'consumption')
    
    return consumption_df

def biofuel_simplified_modelling_methods_handler(production_df, consumption_df, input_data_dict, model_df_clean_wide):
    """ handler for the simplified modelling methods for biofuels. This will be used to set the economy to continue with the same ratio of production to consumpt or to satisfy all demand with domestic production (or RAMP)."""
    simplified_economy_fuels = input_data_dict['simplified_economy_fuels']
    economy = production_df['economy'].unique()[0]
    simplified_economy_fuels = simplified_economy_fuels[simplified_economy_fuels['economy'] == economy]
<<<<<<< HEAD
    
=======
    # breakpoint()
>>>>>>> d17335e41b94593dde28c4249295c1c398b475b6
    if simplified_economy_fuels.shape[0] == 0:
        return production_df, consumption_df
    
    for fuel in simplified_economy_fuels['fuel'].unique():
<<<<<<< HEAD
=======
        
>>>>>>> d17335e41b94593dde28c4249295c1c398b475b6
        fuel_production_df = production_df[production_df['subfuels'] == fuel].copy().reset_index(drop = True)
        fuel_consumption_df = consumption_df[consumption_df['subfuels'] == fuel].copy().reset_index(drop = True)
        method = simplified_economy_fuels[simplified_economy_fuels['fuel'] == fuel]['method'].values[0]
        
        if len(simplified_economy_fuels[simplified_economy_fuels['fuel'] == fuel]) > 1:
            raise Exception(f'More than one method for {fuel} in simplified_economy_fuels')
        
        if method == 'keep_same_ratio_of_production_to_consumption':
            fuel_production_df, fuel_consumption_df = keep_same_ratio_of_production_to_consumption(economy, model_df_clean_wide,fuel_production_df, fuel_consumption_df,  input_data_dict)
        elif method == 'satisfy_all_demand_with_domestic_production':
            fuel_production_df, fuel_consumption_df = satisfy_all_demand_with_domestic_production(economy, model_df_clean_wide, fuel_production_df, fuel_consumption_df, input_data_dict)
<<<<<<< HEAD
=======
        elif method == 'do_nothing':
            pass
>>>>>>> d17335e41b94593dde28c4249295c1c398b475b6
        elif 'satisfy_all_demand_with_domestic_production_RAMP' in method:
            years_to_ramp_over = re.search(r'\d', method)
            if not years_to_ramp_over:
                years_to_ramp_over = 10
            fuel_production_df, fuel_consumption_df =satisfy_all_demand_with_domestic_production(economy, model_df_clean_wide, fuel_production_df, fuel_consumption_df, input_data_dict, RAMP=True, years_to_ramp_over = years_to_ramp_over)
<<<<<<< HEAD
            
=======

>>>>>>> d17335e41b94593dde28c4249295c1c398b475b6
        else:
            raise Exception(f'{method} is not a valid method for biofuel simplified modelling')
        
        production_df = production_df[production_df['subfuels'] != fuel].copy().reset_index(drop = True)
        production_df = pd.concat([production_df, fuel_production_df], ignore_index = True)
        consumption_df = consumption_df[consumption_df['subfuels'] != fuel].copy().reset_index(drop = True)
        consumption_df = pd.concat([consumption_df, fuel_consumption_df], ignore_index = True)
        
    return production_df, consumption_df
    
def keep_same_ratio_of_production_to_consumption(economy, model_df_clean_wide,fuel_production_df, fuel_consumption_df,  input_data_dict):
<<<<<<< HEAD
    """This is for where we want the economy, for a certain fuel, to continue with the same ratio of production to consumption. That will also lead to the same ratio of imports and exports. This is useful for economies where we dont have capacity data and just want to push it out at the same rate it was."""
=======
    """This is for where we want the economy, for a certain fuel, to continue with the same ratio of production to consumption. That will also lead to the same ratio of imports and exports. This is useful for economies where we dont have capacity data and just want to push it out at the same rate it was.
    
    If working with a series that has already had capcaity data entered in, it WILL overwrite that data to make the ratio of production to consumption the same for all years. """
>>>>>>> d17335e41b94593dde28c4249295c1c398b475b6
    proj_years = list(range(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR+1, 1))
    historical_years = list(range(EBT_EARLIEST_YEAR, OUTLOOK_BASE_YEAR+1, 1))
    all_years = historical_years + proj_years
    all_years = [year for year in all_years if year in model_df_clean_wide.columns]
<<<<<<< HEAD
=======
    
    #double check that the values for production the same as OUTLOOK_BASE_YEAR for the rest of the years, otherwise it inidcates taht the modeller tried to set the capcity data and this will overwrite that. To be safe we will always throw an error in this case, and ask the mdeller to use the satisfy_all_demand_with_domestic_production method or no method at all instead.
    for scenario in fuel_production_df['scenarios'].unique():
        for subfuel in fuel_production_df['subfuels'].unique():
            if not fuel_production_df[(fuel_production_df['scenarios'] == scenario) & (fuel_production_df['subfuels'] == subfuel) & (fuel_production_df['year'] == OUTLOOK_BASE_YEAR)]['production'].equals(fuel_production_df[(fuel_production_df['scenarios'] == scenario) & (fuel_production_df['subfuels'] == subfuel) &(fuel_production_df['year'] > OUTLOOK_BASE_YEAR)]['production']):
                raise Exception(f'Production values for {subfuel} in {scenario} are not the same for all years. This indicates that capacity data has been entered in. Please use the satisfy_all_demand_with_domestic_production method or no method instead, since the keep_same_ratio_of_production_to_consumption will overwrite the capcity data you had entered in.')
            
            
>>>>>>> d17335e41b94593dde28c4249295c1c398b475b6
    production_fuel_consumption_df = fuel_production_df.merge(fuel_consumption_df, on = ['scenarios', 'economy', 'subfuels', 'year'], how = 'outer', suffixes=('_production', '_consumption'))
    #calculate the ratio
    production_fuel_consumption_df['ratio'] = production_fuel_consumption_df['production'] / production_fuel_consumption_df['consumption']
    #set any nas to 1 
    production_fuel_consumption_df['ratio'] = production_fuel_consumption_df['ratio'].fillna(1)
    #set any infs to 1 too
    production_fuel_consumption_df['ratio'] = production_fuel_consumption_df['ratio'].replace([np.inf, -np.inf], 1)
    
    #set all ratios after the base year to the base year ratio
    for scenario in production_fuel_consumption_df['scenarios'].unique():
        production_fuel_consumption_df.loc[(production_fuel_consumption_df['year'] > OUTLOOK_BASE_YEAR) & (production_fuel_consumption_df['scenarios'] == scenario), 'ratio'] = production_fuel_consumption_df.loc[(production_fuel_consumption_df['year'] == OUTLOOK_BASE_YEAR) & (production_fuel_consumption_df['scenarios'] == scenario), 'ratio'].values[0]
        
    # Now we need to set production to be the ratio of consumption for projected years
    production_fuel_consumption_df.loc[production_fuel_consumption_df['year'] > OUTLOOK_BASE_YEAR, 'production'] = production_fuel_consumption_df['consumption'] * production_fuel_consumption_df['ratio']
    fuel_production_df = production_fuel_consumption_df[['scenarios', 'economy', 'subfuels', 'year', 'production']].copy().reset_index(drop = True)
    return fuel_production_df, fuel_consumption_df
    
    
    
    
def satisfy_all_demand_with_domestic_production(economy, model_df_clean_wide, fuel_production_df, fuel_consumption_df, input_data_dict, years_to_ramp_over = 0, RAMP = False):
    """This is for where we want the economy, for a certain fuel, to satisfy all demand with domestic production. This is useful for economies where we dont have capacity data and just want them to be self sufficient. This will mean that the imports and exports will be 0. 
    
<<<<<<< HEAD
    There is also an option to ramp it up over time to 100% of demand. That is so that we dont see a sudden jump in production in the future."""
=======
    There is also an option to ramp it up over time to 100% of demand. That is so that we dont see a sudden jump in production in the future.
    
    If working with a series that has already had capcaity data entered in, it will just make sure that production is >= consumption for all years. """
>>>>>>> d17335e41b94593dde28c4249295c1c398b475b6
    
    proj_years = list(range(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR+1, 1))
    historical_years = list(range(EBT_EARLIEST_YEAR, OUTLOOK_BASE_YEAR+1, 1))
    all_years = historical_years + proj_years
    all_years = [year for year in all_years if year in model_df_clean_wide.columns]
    production_fuel_consumption_df = fuel_production_df.merge(fuel_consumption_df, on = ['scenarios', 'economy', 'subfuels', 'year'], how = 'outer', suffixes = ('_production', '_consumption'))
    #calculate the ratio
    production_fuel_consumption_df['ratio'] = production_fuel_consumption_df['production'] / production_fuel_consumption_df['consumption']
    #set any nas to 1
    production_fuel_consumption_df['ratio'] = production_fuel_consumption_df['ratio'].fillna(1)
    
    #set any infs to 1 too
    production_fuel_consumption_df['ratio'] = production_fuel_consumption_df['ratio'].replace([np.inf, -np.inf], 1)
    
<<<<<<< HEAD
    #if RAMP then we need to calculate the amount to ramp up by each year until ratio is >=1    (so if its already 1 then we dont need to ramp -  instead just make sure that production is >= consumption)
=======
    #if RAMP then we need to calculate the amount to ramp up by each year until ratio is >=1    (so if its already >=1 then we dont need to ramp -  instead just make sure that production is >= consumption - 'greater than' is important in case the capacity data we may have had supplied is already cuasing production to be greater than consumption, in which case there would need to be exports, which is ok)
>>>>>>> d17335e41b94593dde28c4249295c1c398b475b6
    if RAMP: 
        #grab the raTIO between production and consumption for each fuel type iuun the base year
        years_to_exclude = [OUTLOOK_BASE_YEAR]
        ratio_base_year = production_fuel_consumption_df[production_fuel_consumption_df['year'] == OUTLOOK_BASE_YEAR]
        for scenario in production_fuel_consumption_df['scenarios'].unique():
            ratio = production_fuel_consumption_df[(production_fuel_consumption_df['scenarios'] == scenario)]['ratio'].values[0]
            if ratio < 1:
                #calculate the amount to ramp up by each year until ratio is >=1
                annual_ramp_up_amount = ratio / years_to_ramp_over
                for year in range(OUTLOOK_BASE_YEAR+1, OUTLOOK_BASE_YEAR+years_to_ramp_over+1, 1):
                    years_to_exclude.append(year)
                    ratio += annual_ramp_up_amount
                    if ratio > 1:
                        breakpoint()
                    production_fuel_consumption_df.loc[(production_fuel_consumption_df['year'] == year), 'production'] = production_fuel_consumption_df.loc[(production_fuel_consumption_df['year'] == year), 'consumption'] * ratio
            #make sure that production is >= consumption for all years except the years_to_exclude
            production_fuel_consumption_df.loc[(production_fuel_consumption_df['production'] < production_fuel_consumption_df['consumption']) & (~production_fuel_consumption_df['year'].isin(years_to_exclude)) & (production_fuel_consumption_df['scenarios'] == scenario) & (production_fuel_consumption_df['ratio'] < 1), 'production'] = production_fuel_consumption_df.loc[(production_fuel_consumption_df['production'] < production_fuel_consumption_df['consumption']) & (~production_fuel_consumption_df['year'].isin(years_to_exclude)) & (production_fuel_consumption_df['scenarios'] == scenario) & (production_fuel_consumption_df['ratio'] < 1), 'consumption']
            
<<<<<<< HEAD
    #now we need to set any production that is less than consumption to consumption for years after the base year
=======
    #now we need to set any production that is less than consumption to consumption for years after the base year (- 'less than' is important in case the capacity data we may have had supplied is already cuasing production to be greater than consumption, in which case there would need to be exports, which is ok)
>>>>>>> d17335e41b94593dde28c4249295c1c398b475b6
    production_fuel_consumption_df.loc[(production_fuel_consumption_df['year'] > OUTLOOK_BASE_YEAR) & (production_fuel_consumption_df['production'] < production_fuel_consumption_df['consumption']), 'production'] = production_fuel_consumption_df.loc[(production_fuel_consumption_df['year'] > OUTLOOK_BASE_YEAR) & (production_fuel_consumption_df['production'] < production_fuel_consumption_df['consumption']), 'consumption']
        
    fuel_production_df = production_fuel_consumption_df[['scenarios', 'economy', 'subfuels', 'year', 'production']].copy().reset_index(drop = True)
    return fuel_production_df, fuel_consumption_df


def calculate_exports_imports(model_df_clean_wide, consumption_df, production_df):
    
    #merge consumption and production data
    final_df = consumption_df.merge(production_df, on = ['scenarios', 'economy', 'subfuels', 'year'], how = 'outer')
    #find any nas in either production or consumption and let user know, we should at least have 0's for these
    if final_df['consumption'].isnull().sum() > 0:
        print('There are {} NaNs in the consumption data'.format(final_df['consumption'].isnull().sum()))
        breakpoint()
        raise Exception('There are NaNs in the consumption data')
    if final_df['production'].isnull().sum() > 0:
        print('There are {} NaNs in the production data'.format(final_df['production'].isnull().sum()))
        breakpoint()
        raise Exception('There are NaNs in the production data')
    #calculate imports and exports. this is jsut the difference between consumption and production, and if negative, it is an export if positive it is an import
    final_df['02_imports'] = np.where(final_df['consumption'] - final_df['production'] > 0, (final_df['consumption'] - final_df['production']), 0)
    final_df['03_exports'] = np.where(final_df['consumption'] - final_df['production'] < 0, (final_df['consumption'] - final_df['production']), 0)
    
    #rename 01_production:
    final_df.rename(columns = {'production':'01_production'}, inplace = True)
    
    # #we now have no need for the consumption column
    # final_df.drop(columns = ['consumption'], inplace = True)
    
    #not sure if we need to fill in biofuels_processing in transformation so leave it for now.
    
    
    #melt the sectors (01_production, 02_imports, 03_exports) so they are in one column and then pivot the years so they are in columns:
    final_df = final_df.melt(id_vars = ['scenarios', 'economy', 'subfuels', 'year'], value_vars = ['consumption', '01_production', '02_imports', '03_exports'], var_name = 'sectors', value_name = 'energy_pj')
    year_cols = final_df['year'].unique().tolist()
    #pivot years
    final_df = final_df.pivot_table(index = ['scenarios', 'economy', 'subfuels', 'sectors'], columns = 'year', values = 'energy_pj').reset_index()
    
    #fill in missing cols from model_df_clean_wide:
    model_df_clean_wide_index_cols = model_df_clean_wide[['scenarios', 'economy', 'sectors', 'sub1sectors', 'sub2sectors', 'sub3sectors', 'sub4sectors', 'fuels', 'subfuels']].drop_duplicates().copy().reset_index(drop = True)
    final_df = model_df_clean_wide_index_cols.merge(final_df, on = ['scenarios', 'economy', 'sectors', 'subfuels'], how = 'right')
    #check for any missing values, this is where we may have created a category that wasnt in the original model_df_clean_wide
    #exceopt ignore where sectors==consumption as this will be missing
    if final_df[final_df['sectors'] != 'consumption'].isnull().sum().sum() > 0:
        print('There are {} NaNs in the final_df'.format(final_df[final_df['sectors'] != 'consumption'].isnull().sum().sum()))
        breakpoint()
        raise Exception('There are NaNs in the final_df')
    return final_df

def plot_biofuels_data(final_refining_df, economy):
    """
    Create and display Plotly Express charts based on the final_refining_df data.
    """
    
    import plotly.express as px
    # Filter data for the specific economy and melt for visualization.
    df = final_refining_df[final_refining_df['economy'] == economy].copy()
    df_melted = df.melt(id_vars=['scenarios', 'economy', 'sectors', 'sub1sectors', 'sub2sectors', 'sub3sectors', 'sub4sectors', 'fuels', 'subfuels'], var_name='year', value_name='energy_pj')

    # Line chart to show the production, imports, and exports over the years by fuel type.
    fig_production = px.line(
        df_melted,
        x='year',
        y='energy_pj',
        color='sectors',
        line_dash='sectors',
        facet_row='scenarios',
        facet_col='subfuels',
        title=f'Biofuels Production, Imports, and Exports Over Time for {economy}',
        labels={'year': 'Year', 'energy_pj': 'Energy (PJ)', 'sectors': 'Supply Components', 'subfuels': 'Fuel Type'}
    )

    # Make the y-axes independent
    fig_production.update_yaxes(matches=None, showticklabels=True)
    
    #write html
    fig_production.write_html(f'./plotting_output/biofuels/{economy}_biofuels_line.html')
    
    # Bar chart to compare the production, imports, and exports by sector and year
    fig_bar = px.bar(
        df_melted.loc[df_melted['sectors'] != 'consumption'],
        x='year',
        y='energy_pj',
        color='sectors',
        barmode='relative',
        facet_col='scenarios',
        facet_row='subfuels',
        title=f'Biofuels Production, Imports, and Exports Comparison for {economy}',
        labels={'year': 'Year', 'energy_pj': 'Energy (PJ)', 'sectors': 'Supply Components'}
    )

    # Make the y-axes independent
    fig_bar.update_yaxes(matches=None, showticklabels=True)
    #write html
    fig_bar.write_html(f'./plotting_output/biofuels/{economy}_biofuels_stacked_bars.html')
    
    #and now create one which has net imports as a line, with prodcuton and conshunption in the chart too
    df_net_imports = df_melted.copy()
    #pivot the sectors so they are in columns
    df_net_imports = df_net_imports[['scenarios', 'economy', 'subfuels', 'year', 'sectors', 'energy_pj']].pivot(index = ['scenarios', 'economy', 'subfuels', 'year'], columns = 'sectors', values = 'energy_pj').reset_index()
    
    df_net_imports['net_imports'] = df_net_imports['02_imports'] + df_net_imports['03_exports']
    df_net_imports.drop(columns = ['02_imports', '03_exports'], inplace = True)
    df_net_imports = df_net_imports.melt(id_vars=['scenarios', 'economy', 'subfuels', 'year'], var_name='sectors', value_name='energy_pj')
    fig_net_imports = px.line(
        df_net_imports,
        x='year',
        y='energy_pj',
        color='sectors',
        line_dash='sectors',
        facet_row='scenarios',
        facet_col='subfuels',
        title=f'Biofuels Production, Consumption and Net Imports Over Time for {economy} - negative imports are exports',
        labels={'year': 'Year', 'energy_pj': 'Energy (PJ)', 'sectors': 'Supply Components', 'subfuels': 'Fuel Type'}
    )

    # Make the y-axes independent
    fig_net_imports.update_yaxes(matches=None, showticklabels=True)
    #write html
    fig_net_imports.write_html(f'./plotting_output/biofuels/{economy}_biofuels_net_imports_line.html')
    

def save_results(final_df, detailed_capacity_df, capacity_df, economy):
    # Save location
    save_location = './results/supply_components/04_biofuels/{}/'.format(economy)
    #drop 'consumption' from final_df's sectors column
    final_df = final_df.loc[final_df['sectors'] != 'consumption']
    
    if not os.path.isdir(save_location):
        os.makedirs(save_location)
        
    for scenario in final_df['scenarios'].unique():
        supply_df = final_df[final_df['scenarios'] == scenario].copy().reset_index(drop = True)
        #save to a folder to keep copies of the results
        supply_df.to_csv(save_location + economy + '_biofuels_' + scenario + '_' + timestamp + '.csv', index = False)     
        #and save them to modelled_data folder too. but only after removing the latest version of the file
        for file in os.listdir(f'./data/modelled_data/{economy}/'):
            if re.search(economy + '_biofuels_' + scenario, file):
                os.remove(f'./data/modelled_data/{economy}/' + file)
        
        supply_df.to_csv(f'./data/modelled_data/{economy}/' + economy + '_biofuels_' + scenario + '_' + timestamp + '.csv', index = False)
    #and save detailed_capacity_df to the results folder but only after removing the latest version of the file
    for file in os.listdir('./results/supply_components/04_biofuels/{}/'.format(economy)):
        if re.search(economy + '_biofuels_capacity_additions_detailed', file):
            if os.path.exists('./results/supply_components/04_biofuels/{}/'.format(economy) + file):
                os.remove('./results/supply_components/04_biofuels/{}/'.format(economy) + file)
        if re.search(economy + '_biofuels_capacity', file):
            if os.path.exists('./results/supply_components/04_biofuels/{}/'.format(economy) + file):
                os.remove('./results/supply_components/04_biofuels/{}/'.format(economy) + file)
       
    detailed_capacity_df.to_csv(save_location + economy + '_biofuels_capacity_additions_detailed_' + timestamp + '.csv', index = False)
    capacity_df.to_csv(save_location + economy + '_biofuels_capacity_' + timestamp + '.csv', index = False)

