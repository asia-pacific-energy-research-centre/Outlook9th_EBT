import pandas as pd
import numpy as np
import os
import glob
from datetime import datetime
from utility_functions import *
import warnings
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def label_subtotals(results_layout_df, shared_categories):  
    def label_subtotals_for_sub_col(df, sub_col):
        """
        This fucntion is run for each sub_col. It identifies and labels subtotal rows within the dataframe, based on the current subcol 
        
        If this row is not a subtotal for this sub_col then it wont change anythign, but if it is a subtotal for this sub_col then it will label it as a subtotal.
        
        (so, if the row is identified as a subtotal, then it will be labelled as a subtotal, but technically this means that the sub_col will be 'x' (see x_mask varaible), since it is a subtotal of all values in that sub_col, but for the category one level higher than that sub_col).

        This function goes through the dataframe to find subtotal rows based on the presence of 'x' in specific columns.
        
        It distinguishes between definite non-subtotals (where tehre is no data to aggregate) and potential subtotals (where there's a mix of detailed and aggregated data in that group). It then labels the potential subtotals accordingly.
        
        A similar thing to the x's has been added for the presence of 0's and nans. Note that a value is  definitely not a subtotal if EITHER of the conditions are met.
        
        Parameters:
        - df (DataFrame): The dataframe to process.
        - sub_col (str): The name of the column to check for 'x' which indicates a subtotal.
        - totals (list): A list of values in the 'fuels' column that should be considered as total values and not subtotals.

        Returns:
        - DataFrame: The original dataframe with an additional column indicating whether a row is a subtotal.
        """
        #############################
        # the following section finds where all the values in a group for the sub_col are x. These are not subtotals since there must be no more specific values than them if there arent any non x values at all for this group in this sub_col. dont label them at all (since they will default to is_subtotal=False).
        # Create a mask for 'x' values in the sub_col 
        x_mask = (df[sub_col] == 'x')

        # Group by all columns except 'value' and sub_col and check if all values in sub_col are 'x' for that group 
        grouped = df.loc[x_mask, [col for col in df.columns if col not in ['value', sub_col, 'is_subtotal']]].groupby([col for col in df.columns if col not in ['value', sub_col, 'is_subtotal']]).size().reset_index(name='count')

        # Merge the original df with the grouped data to get the count of 'x' for each group
        merged = df.merge(grouped, on=[col for col in df.columns if col not in ['value', sub_col, 'is_subtotal']], how='left')

        # Create a mask where the count is equal to the size of the group, indicating all values in sub_col were 'x' for that group.
        non_subtotal_mask = merged['count'] == df.groupby([col for col in df.columns if col not in ['value', sub_col, 'is_subtotal']]).transform('size').reset_index(drop=True)
        
        ############################# 
        #if more than one value are not zero/nan for this group, then it could be a subtotal, if not, its a definite non-subtotal since its the most specific data we have for this group.
        value_mask = (abs(df['value'])> 0)
        # Group by all columns except 'value' and sub_col and check how many values are >0 or <0 for that group
        grouped = df.loc[value_mask, [col for col in df.columns if col not in ['value', sub_col, 'is_subtotal']]].groupby([col for col in df.columns if col not in ['value', sub_col, 'is_subtotal']]).size().reset_index(name='count')

        # Merge the original df with the grouped data to get the count of 'x' for each group
        merged = df.merge(grouped, on=[col for col in df.columns if col not in ['value', sub_col, 'is_subtotal']], how='left')

        #fill nas in merged with True, as this is where no values are >0 or <0 for this group
        merged['count'] = merged['count'].fillna(True)
        
        # Create a mask where the count of values in the group that are >0 or <0 is <= 1, indicating that its not a subtotal 
        non_subtotal_mask2 = merged['count'] <= 1
        #############################
        
        # df.reset_index(drop=True, inplace=True)
        df['non_subtotal_mask'] = non_subtotal_mask
        df['non_subtotal_mask2'] = non_subtotal_mask2
        #separate where all the values in the group are x. these are not subtotals since there are no more specific values than them. dont label them at all.
        
        df_definitely_not_subtotals = df[non_subtotal_mask2 | non_subtotal_mask].copy()
        df_maybe_subtotals = df[~(non_subtotal_mask2 | non_subtotal_mask)].copy()
        conditions = {
            'sub_col_is_x': df_maybe_subtotals[sub_col] == 'x'
        }
        # Combine conditions for subtotals
        conditions['is_subtotal'] =  conditions['sub_col_is_x'] 
        
        # Apply the conditions to label subtotal rows. but only label if it is a subtotal. dont label with false if it is not a subtotal. this is because it could overwrite where we previously labelled it as a subtotal.
        df_maybe_subtotals['is_subtotal'] = np.where(conditions['is_subtotal'], True, df_maybe_subtotals['is_subtotal'])
        # Concatenate the two DataFrames back together
        df = pd.concat([df_definitely_not_subtotals, df_maybe_subtotals], ignore_index=True).reset_index(drop=True)
        
        #drop non_subtotal_mask and non_subtotal_mask2
        df.drop(columns=['non_subtotal_mask', 'non_subtotal_mask2'], inplace=True)
        
        return df
    #####################################################################
    #SUBFUNCTION
    ######################################################################
    if 'is_subtotal' in results_layout_df.columns:
        results_layout_df = results_layout_df.drop(columns=['is_subtotal'])
    # Melt the DataFrame
    df_melted = results_layout_df.melt(id_vars=shared_categories, var_name='year', value_name='value')
    
    #Double check for duplicates:
    dupes = df_melted[df_melted.duplicated(subset=shared_categories+['year'], keep=False)]
    if dupes.shape[0] > 0:
        print('Found duplicates in subtotal input')
    
    #make sure year is int
    df_melted['year'] = df_melted['year'].astype(int)
    
    #remove year col and sum value col. this will allow us to more easily identify if a value is standalone or a sum of others. 
    df_melted_sum = df_melted.groupby([col for col in df_melted.columns if col not in ['year', 'value']]).sum(numeric_only=True).reset_index()
    df_melted_sum.drop(columns=['year'], inplace=True)

    #set is_subtotal to False. It'll be set to True, eventually, if it is a subtotal
    df_melted_sum['is_subtotal'] = False
    for sub_col in ['subfuels', 'sub4sectors', 'sub3sectors', 'sub2sectors', 'sub1sectors']:
        df_melted_sum = label_subtotals_for_sub_col(df_melted_sum, sub_col)
    
    #join df_melted_sum, wthout value col to df_melted on [col for col in df_melted.columns if col not in ['value'] to get the is_subtotal col in the original df_melted
    df_melted = pd.merge(df_melted, df_melted_sum[[col for col in df_melted_sum.columns if col not in ['value']]], on=[col for col in df_melted.columns if col not in ['year', 'value']], how='left')
    
    # # Drop rows where 'value' column has NaN
    # df_melted = df_melted.dropna(subset=['value']).copy()
    
    # Group by all columns except 'value' and sum the 'value' column
    df_melted = df_melted.groupby([col for col in df_melted.columns if col != 'value']).sum().reset_index().copy()
    
    duplicates = df_melted[df_melted.duplicated(subset=shared_categories + ['year'], keep=False)]
    if not duplicates.empty:
        print("Duplicates found:", duplicates)
        duplicates.to_csv('data/temp/error_checking/duplicates_in_tfc.csv', index=False)
    
    df_melted = df_melted.pivot(index=[col for col in df_melted.columns if col not in ['year', 'value']], columns='year', values='value').reset_index()
    #make sure year cols is str
    df_melted.columns = [str(col) for col in df_melted.columns]
    return df_melted

def calculate_subtotals(df, shared_categories, DATAFRAME_ORIGIN):
    """
    Aggregates subtotals for each combination of categories. Then drops any of these calcalted subtotals that are already in the data. This is because the data that is already in there is actually the most specific data for each group, so we should keep it, rather than the subtotal.
    
    Please note that this function requires labelling of subtotals to have already been done. This is because it will look for any subtotals that are already in the data and remove them. If there arent any subtotals in the data, then this function will eventually throw an error.
    
    
    Args:
        df (pd.DataFrame): The input data frame.
        shared_categories (list): List of shared category column names.

    Returns:
        pd.DataFrame: DataFrame with aggregated subtotals.
    """
    #drop any subtotals already in the data, as we will recalcualte them soon. 
    df_no_subtotals = df[df['is_subtotal'] == False].copy()
    df_no_subtotals = df_no_subtotals.drop(columns=['is_subtotal']).copy()
    ############################
    def calculate_subtotal_for_columns(melted_df, cols_to_sum):
        #gruop by the shared categories except the ones we are summing (in cols_to_sum) and sum the values. By doing this on each combination of the shared cateorgires, starting from the most specific ones, we can create subtotals for each combination of the shared categories.        
        group_cols = [col for col in melted_df.columns if col not in ['value'] and col not in cols_to_sum]
        group_cols_no_year = [col for col in group_cols if col != 'year']
        agg_df = melted_df.copy()
        
        #ignore where all of the cols in cols_to_sum are 'x'. This is where  the data is at its most detailed level already. We will ignore these rows and tehrefor not create a subtotal for them.
        agg_df = agg_df[~(agg_df[cols_to_sum] == 'x').all(axis=1)].copy()#but what this will do is create duplcates?i think this is the falt of mixing subfuels and subsectors as they are independt. need to trawt them separately
        
        
        #note i added the below code on 4/2/2025 as it seemed it was better tan the ither code
        #and ignore where only one, or less, values are non zero/nan for this group (group_cols), as this is also the most detailed level of data (or there is at least no data to sum up to a subtotal)and so shoudlnt have its values replaced with x'sand be summed up. #please note there is a very small chance this causes issues. i think its if we have a row that has a unique subfuels and sector and is only non zero for one year. This will cause it to be dropped from the df and not included in the subtotal. But this is very unlikely to happen, so we will ignore it for now.
        # breakpoint()
        agg_df_more_than_one_non_zero_no_years = melted_df[group_cols_no_year +['value']].groupby(group_cols_no_year, as_index=False).sum().reset_index().copy()
        agg_df_more_than_one_non_zero_no_years = agg_df_more_than_one_non_zero_no_years.loc[(agg_df_more_than_one_non_zero_no_years['value'] != 0) & (agg_df_more_than_one_non_zero_no_years['value'].notnull())].copy()
        agg_df_more_than_one_non_zero_no_years = agg_df_more_than_one_non_zero_no_years.groupby(group_cols_no_year, as_index=False)['value'].count().reset_index().copy()
        
        agg_df_more_than_one_non_zero_no_years = agg_df_more_than_one_non_zero_no_years[agg_df_more_than_one_non_zero_no_years['value'] != 0].copy()
        agg_df = agg_df.merge(agg_df_more_than_one_non_zero_no_years[group_cols_no_year], on=group_cols_no_year, how='left', indicator=True)
        agg_df = agg_df[agg_df['_merge'] == 'both'].copy()
        agg_df.drop(columns=['_merge'], inplace=True)       
        
        #identify if 15_03_01_passenger or 15_03_02_freight are in sub2sectors, if so

        # # #join that to the melted df and remove where there is no match. This will stop the effect of years? nah?
        
        #note i removed the below code on 4/2/2025 as it seemed the above code was better
        # agg_df_more_than_one_non_zero_no_years = melted_df.loc[(melted_df['value'] != 0) & (melted_df['value'].notnull())].copy()
        # agg_df_more_than_one_non_zero = melted_df.loc[(melted_df['value'] != 0) & (melted_df['value'].notnull())].copy()
        # agg_df_more_than_one_non_zero = agg_df_more_than_one_non_zero.groupby(group_cols_no_year, as_index=False)['value'].count().reset_index().copy()
        # # agg_df_one_non_zero = agg_df_more_than_one_non_zero.loc[agg_df_more_than_one_non_zero['value'] == 1].copy()#unlikely. maybe we should be searching for only 1 or more non zeros? #inore me if you unvcomment the others, its not useful.. just conisdering if we can do ==1 instead of >1
        # agg_df_more_than_one_non_zero = agg_df_more_than_one_non_zero[agg_df_more_than_one_non_zero['value'] > 1].copy()
        # agg_df = agg_df.merge(agg_df_more_than_one_non_zero[group_cols_no_year], on=group_cols_no_year, how='left', indicator=True)
        # agg_df = agg_df[agg_df['_merge'] == 'both'].copy()
        # agg_df.drop(columns=['_merge'], inplace=True)       
        
        # agg_df = agg_df[~(agg_df[cols_to_sum] == 'x').any(axis=1)].copy()#i think this is causeing us to drop allpossible rows to gruop by. resultuiing in no new subtotals being created. TODO: investigate this.
        agg_df = agg_df.groupby(group_cols, as_index=False)['value'].sum().copy()
        for omit_col in cols_to_sum:
            agg_df[omit_col] = 'x'
            
        #dont label as subttoal yet

        return agg_df
    ############################
    
    # Create a list to store each subset result
    subtotalled_results = pd.DataFrame()

    # Define different subsets of columns to find subtotals for. Should be every combination of cols when you start from the msot specific and add one level of detail each time. but doesnt include the least specifc categories since those cannot be subtotaled to a moer detailed level. 
    #if the columns change from the layout then this will need to be updated.
    sets_of_cols_to_sum = [
        ['sub4sectors'],
        ['sub3sectors', 'sub4sectors'],
        ['sub2sectors', 'sub3sectors', 'sub4sectors'],
        ['sub1sectors', 'sub2sectors', 'sub3sectors', 'sub4sectors']
    ]
    melted_df = df_no_subtotals.melt(id_vars=shared_categories,
                            value_vars=[col for col in df_no_subtotals.columns if str(col).isnumeric()],
                            var_name='year',
                            value_name='value')
    melted_df_with_subtotals = df.melt(id_vars=shared_categories+['is_subtotal'],
                            value_vars=[col for col in df.columns if str(col).isnumeric()],
                            var_name='year',
                            value_name='value')
    
    # #sum up vlaues so we can check for where there are more than one na/0 in the value col for the group_cols. 
    # melted_df = melted_df.groupby(shared_categories, as_index=False)['value'].sum().reset_index().copy()  
    #check for duplicates:
    duplicates = melted_df[melted_df.duplicated(subset=shared_categories+['year'], keep=False)]
    if duplicates.shape[0] > 0:
        duplicates.to_csv('data/temp/error_checking/duplicates_in_subtotaled_df.csv', index=False)
        print("WARNING: There are duplicates in the subtotaled DataFrame.")
        breakpoint()  
    
    # Filter out rows with '12_x_other_solar' in 'subfuels'
    # melted_df = melted_df[melted_df['subfuels'] != '12_x_other_solar']

    # Save the dataframe before subtotal calculations
    # melted_df.to_csv('melted_df_before_subtotals.csv', index=False)

    # Process the DataFrame with each cols_to_sum combination so you get a subtotal calculated for every level of detail.
    
    # if DATAFRAME_ORIGIN =='layout':#to do with hkc inudstry
    #     #we're going to filter for only the industry sector and petrolum pridycts aso that investigating isssues is easyuer.
    #     melted_df = melted_df[(melted_df['sectors'] == '14_industry_sector') & (melted_df['fuels'] == '07_petroleum_products')].copy()
        
    for cols_to_sum in sets_of_cols_to_sum:
        # if DATAFRAME_ORIGIN =='layout':#to do with hkc inudstry
        #     breakpoint()
        # if cols_to_sum == ['sub3sectors', 'sub4sectors']:
        #     breakpoint()#why do we lose the jet fuel row here?
        subtotalled_results = pd.concat([subtotalled_results,calculate_subtotal_for_columns(melted_df, cols_to_sum)], ignore_index=True)
    # if DATAFRAME_ORIGIN =='layout':#to do with hkc inudstry
    #     breakpoint()
    #then run it to calauclte a subtotl of each group, including the new subtotals for subsectors, where the subfuel is x. 
    subtotalled_results = pd.concat([subtotalled_results,calculate_subtotal_for_columns(pd.concat([subtotalled_results,melted_df]), ['subfuels'])], ignore_index=True)
    
    # Fill 'x' for the aggregated levels as they will just be nas
    for col in sets_of_cols_to_sum[-1]:
        # Save the dataframe to CSV for inspection
        # subtotalled_results.to_csv('subtotalled_results.csv', index=False)

        #check for nas
        if subtotalled_results[col].isna().any():
            #fill nas with x
            breakpoint()
            duplicates.to_csv('data/temp/error_checking/duplicates_in_subtotals.csv', index=False)
            raise Exception("WARNING: There are nas in the subtotaled DataFrame.")#uplifted to exception as this is unexpected and should be investigated.
            # subtotalled_results[col] = subtotalled_results[col].fillna('x').copy()

    ############################
    
    #check for duplicates in these new subtotals based on all the cols except value.. if there are duplicates then perhaphs something went wrong with the subtotaling. But if the values are all 0 then just drop them.
    #Note: We are taking a risk of not catching potential errors if we assume we can drop subtotals that are duplciates and = 0...but if we dont do this it is quite complicated to work find the error anyway. It should only happen if there is not specific enough data in the layout file, so 0's aregetting subtotaled, and then being foudn asduplicates beside what is actually the most specific data for that group. So we will drop these by ideniftying if we are working on the layout file, otherwise throw an error.
    duplicates = subtotalled_results[subtotalled_results.duplicated(subset=[col for col in subtotalled_results.columns if col not in ['value']], keep=False)]
    if duplicates.shape[0] > 0:
        duplicates_non_zero = duplicates[duplicates['value'] != 0].copy()
        duplicates_non_zero = duplicates_non_zero[duplicates_non_zero.duplicated(subset=[col for col in duplicates_non_zero.columns if col not in ['value']], keep=False)]
        if duplicates_non_zero.shape[0] > 0 or DATAFRAME_ORIGIN != 'layout':
            #sort them
            duplicates.sort_values(by=[col for col in duplicates.columns if col not in ['value']], inplace=True)
            if duplicates.shape[0] > 0:
                print(duplicates)
                breakpoint()
                duplicates.to_csv('data/temp/error_checking/duplicates_in_subtotaled_dataframe.csv', index=False)
                raise Exception("There are duplicates in the subtotaled DataFrame.")
        else:
            #sort by value and keep the last duplicate always since it will not be 0.
            subtotalled_results = subtotalled_results.sort_values(by=['value'], ascending=True).drop_duplicates(subset=[col for col in subtotalled_results.columns if col not in ['value']], keep='last').copy()
    
    #now we've dropped all duplicates except those that might be caused by having different origins, we will sum up all values by all cols except origin. this will remove any duplicates that are caused by having different origins.
    if 'origin' in subtotalled_results.columns:
        subtotalled_results = subtotalled_results.groupby([col for col in subtotalled_results.columns if col not in ['value', 'origin']], as_index=False)['value'].sum().copy()
        shared_categories.remove('origin')
    ###################
    
    #now merge with the original df with subtotals. Fristly, we will idenitfy where subtotals dont calacualte to the same vlaue. If the input is a results df then assume its an error. If the input is a layout df then assume that what was originally labelled as a subtotal is in fact a value that wasnt able to be disaggregated. For these values, keep the original, dont calcualte the subtotal and relabel it as not a subtotal.
    # if DATAFRAME_ORIGIN =='layout':#to do with hkc inudstry
    #     breakpoint() #is it here? layout_df.loc[(layout_df['sectors'] == '14_industry_sector') & (layout_df['sub1sectors'] == 'x') & (layout_df['fuels'] == '07_petroleum_products') & (layout_df['subfuels'] == 'x')]
    EPSILON_PERCENT = 0.01
    merged_data = melted_df_with_subtotals.merge(subtotalled_results, on=shared_categories+['year'], how='outer', suffixes=('', '_new_subtotal'), indicator=True)
    #check where the values are different:
    merged_data['value_diff_pct'] = abs((merged_data['value'] - merged_data['value_new_subtotal']) / merged_data['value'])
    subtotals_with_different_values = merged_data[(merged_data['_merge'] == 'both') & (merged_data['value_diff_pct'] > EPSILON_PERCENT) & (merged_data['is_subtotal'] == True)].copy()
    
    ###add some exceptions were already dealt with:
    #drop any 14_industry_sector in sectors if the DATAFRAME_ORIGIN is results:
    if DATAFRAME_ORIGIN == 'results':
        subtotals_with_different_values = subtotals_with_different_values[subtotals_with_different_values['sectors'] != '14_industry_sector'].copy() 
    #drop buildings electricity in fuels if the DATAFRAME_ORIGIN is results:
    if DATAFRAME_ORIGIN == 'results':
        subtotals_with_different_values = subtotals_with_different_values[(subtotals_with_different_values['sub1sectors'] != '16_01_buildings') & (subtotals_with_different_values['fuels'] != '17_electricity')].copy()  
        
    if DATAFRAME_ORIGIN == 'results':
        if subtotals_with_different_values.loc[((subtotals_with_different_values['subfuels'] == 'x') & (subtotals_with_different_values['fuels'] == '16_others') | (subtotals_with_different_values['subfuels'] == 'x') & (subtotals_with_different_values['fuels'] == '15_solid_biomass')) & (subtotals_with_different_values['sectors'] == '01_production')].shape[0] > 0:
            breakpoint()
            raise Exception('We are potentially going to need manually_remove_subtotals_for_15_16_production_in_results however this would bring unexpected side effects for the visualisation code. try to avoid using this if possible. - note i think this is outdated after changing things to be based off 16_others_unallocated and so on')
            # merged_data, subtotals_with_different_values = manually_remove_subtotals_for_15_16_production_in_results(merged_data, subtotals_with_different_values)
    #if we still have subtotals with different values then we will throw an error.          
    ###
    if subtotals_with_different_values.shape[0] > 0 and DATAFRAME_ORIGIN == 'results':
        breakpoint()
        #save the data
        subtotals_with_different_values.to_csv(f'data/temp/error_checking/subtotals_with_different_values.csv', index=False)
        raise Exception('{} subtotals have different values in the original and subtotalled data. This is likely a mistake on the side of the modeller'.format(subtotals_with_different_values.shape[0]))
    elif subtotals_with_different_values.shape[0] > 0 and DATAFRAME_ORIGIN == 'layout':                
        #relable the subtotals as not subtotals
        merged_data.loc[(merged_data['_merge'] == 'both') & (merged_data['value_diff_pct'] > EPSILON_PERCENT) & (merged_data['is_subtotal'] == True), 'is_subtotal'] = False
        #but note that this could create rows where there are two rows with the same shared_categories but not year, but one is a subtotal and one is not. so we will need to drop the isntances where is_subtotal = True so taht we dont get duplicates. We will do this later in this function when we pivot the data wide.
    
    # # Then drop where there are any other subtotals that match rows in the origianl data, but they arent subttotals in the origianl data. Since we removed all labelled subtotals at the start of the funciton we must assume that these matching rows are actual data points for their cateogires, and thereofre shouldnt be replaced with a subtotal! So we will keep the original data and remove the subtotalled data.
    # merged_data = melted_df.merge(subtotalled_results, on=shared_categories+['year'], how='outer', suffixes=('', '_new_subtotal'), indicator=True)
    
    values_to_keep_in_original = merged_data[(merged_data['_merge'] == 'both') & (merged_data['is_subtotal'] == False)].copy()
    values_only_in_original = merged_data[(merged_data['_merge'] == 'left_only')].copy()
    new_subtotalled_values = merged_data[(merged_data['_merge'] == 'right_only')].copy()
    subtotal_values_in_both= merged_data[(merged_data['_merge'] == 'both') & (merged_data['is_subtotal'] == True)].copy()
    manual_subtotal_values = merged_data[(merged_data['_merge'] == 'manual_subtotal')].copy()
    
    values_to_keep_in_original['value'] = values_to_keep_in_original['value']
    values_only_in_original['value'] = values_only_in_original['value']
    new_subtotalled_values['value'] = new_subtotalled_values['value_new_subtotal']
    subtotal_values_in_both['value'] = subtotal_values_in_both['value']
    manual_subtotal_values ['value'] = manual_subtotal_values['value']
    
    new_subtotalled_values['is_subtotal'] = True
    
    #concat all together
    final_df = pd.concat([values_to_keep_in_original, values_only_in_original, new_subtotalled_values, subtotal_values_in_both, manual_subtotal_values], ignore_index=True)
    #drop merge and value_original and value_subtotalled
    final_df.drop(columns=['_merge', 'value_new_subtotal', 'value_diff_pct'], inplace=True)
    
    #TESTING FIND THESE sectors	sub1sectors	sub2sectors	sub3sectors	sub4sectors	fuels	subfuels
    # 14_industry_sector	x	x	x	x	01_coal	x
    # 14_industry_sector	x	x	x	x	01_coal	x
    test = final_df[(final_df['sectors'] == '14_industry_sector') & (final_df['sub1sectors'] == 'x') & (final_df['sub2sectors'] == 'x') & (final_df['sub3sectors'] == 'x') & (final_df['sub4sectors'] == 'x') & (final_df['fuels'] == '08_gas') & (final_df['subfuels'] == 'x') & (final_df['year'].isin([str(EBT_EARLIEST_YEAR), str(OUTLOOK_BASE_YEAR)]))].copy()
    # if test.shape[0] > 0:
        
    #     breakpoint()
    ###################
    #make final_df wide
    #check for duplicates
    duplicates = final_df[final_df.duplicated(subset=shared_categories+['year'], keep=False)]
    #if there are duplicates then save them to a csv so we can check them later and throw an error.
    if duplicates.shape[0] > 0:
        duplicates.to_csv('data/temp/error_checking/duplicates_in_subtotaled_df.csv', index=False)
        breakpoint()
        raise Exception("There are duplicates in the subtotaled DataFrame.")
    final_df_wide = final_df.pivot(index=shared_categories+['is_subtotal'], columns='year', values='value').reset_index()
    ###################
    try:
        #where two rows exist for the same shared_categories but one is a subtotal and one is not, drop the one that is a subtotal.
        same_rows = final_df_wide[final_df_wide.duplicated(subset=shared_categories, keep=False)]
        #find where there are two rows and one is a subtotasl, the other is not:
        to_drop = final_df_wide.loc[same_rows.index.tolist()]
        to_drop = to_drop[to_drop['is_subtotal'] == True].copy()
        to_keep = final_df_wide.loc[same_rows.index.tolist()]
        to_keep = to_keep[to_keep['is_subtotal'] == False].copy()
        #join them on shared_categories and keep only the ones where is_subtotal = False and True. Then drop these rows from final_df_wide
        to_drop = to_drop.merge(to_keep, on=shared_categories, how='left', suffixes=('', '_to_keep'), indicator=True)
        to_drop = to_drop[to_drop['_merge'] == 'both'].copy()
        to_drop = to_drop[shared_categories + ['is_subtotal']].copy()
        final_df_wide = final_df_wide.merge(to_drop, on=shared_categories+['is_subtotal'], how='left', indicator=True)
        final_df_wide = final_df_wide[final_df_wide['_merge'] == 'left_only'].copy()
        final_df_wide.drop(columns=['_merge'], inplace=True)
    except:
        breakpoint()
        print('WARNNG: There was an error when trying to drop subtotals that were duplicates of non-subtotals. This is unexpected and should be investigated. See line 299 in merging_functions.py')
    ###################
    
    # Check again for any duplicates
    duplicates = final_df_wide[final_df_wide.duplicated(subset=[col for col in final_df_wide.columns if col not in ['is_subtotal'] +[year for year in range(EBT_EARLIEST_YEAR, OUTLOOK_BASE_YEAR+1)]], keep=False)]
    if duplicates.shape[0] > 0:
        # print("WARNING: There final_df_wideare duplicates in the subtotaled DataFrame.")
        # print(duplicates)
        breakpoint()
        duplicates.to_csv('data/temp/error_checking/duplicates_in_subtotaled_DF.csv', index=False)
        raise Exception("There are duplicates in the subtotaled DataFrame.") 
    return final_df_wide

def manually_remove_subtotals_for_15_16_production_in_results(merged_data, subtotals_with_different_values):
    # - note i think this is outdated after changing things to be based off 16_others_unallocated and so on
    #this is a bit of a hack. IF THERE ARE clashes between subtotals for production of these fuel types we will manually remove the subtotals for 15_solid_biomass and 16_others in production. This is because we know that we have a mix of results where subfuels is x and subfuels is an actrual subfuel. Note that we previously had the ability to calcualte subtotals here but rteaise it wouldnt work since it would create duplicates when ignoring the subtotal column. i.e. where fuel = 15_solid_biomass and subfuel = x, and subtotal is both true and false.
    subtotals_with_different_values = subtotals_with_different_values[~((subtotals_with_different_values['subfuels'] == 'x') & (subtotals_with_different_values['fuels'] == '16_others') | (subtotals_with_different_values['subfuels'] == 'x') & (subtotals_with_different_values['fuels'] == '15_solid_biomass')) & (subtotals_with_different_values['sectors'] == '01_production')].copy()
    
    new_prod_subtotals = merged_data[(merged_data['sectors'] == '01_production') & (merged_data['fuels'].isin(['16_others', '15_solid_biomass']))].copy()
    #drop nas in value
    new_prod_subtotals = new_prod_subtotals.dropna(subset=['value']).copy()
    # #sum up data by everything excepot subfuel to find the subtotal, and then set subtotal to False for everything else.
    # shared_categories_minus_subfuel = [col for col in shared_categories if col != 'subfuels']
    # subtotals = new_prod_subtotals.groupby(shared_categories_minus_subfuel+['year'], as_index=False)['value'].sum().reset_index().copy()
    #set subfuel to x and is_subtotal to True
    # subtotals['subfuels'] = 'x'
    # subtotals['is_subtotal'] = True
    #thewn in new_prod_subtotals setsubtotal to fasle
    new_prod_subtotals['is_subtotal'] = False
    #then set merge to 'manual_subtotal' so we can identify these rows later
    new_prod_subtotals['_merge'] = 'manual_subtotal'
    # subtotals['_merge'] = 'manual_subtotal'
    #then drop the old rows from merged_data and add the new ones
    merged_data = merged_data[~((merged_data['sectors'] == '01_production') & (merged_data['fuels'].isin(['16_others', '15_solid_biomass'])))].copy()
    merged_data = pd.concat([merged_data, new_prod_subtotals], ignore_index=True)
    return merged_data, subtotals_with_different_values  
def remove_all_zeros(results_df, years_to_keep_in_results):
    #since the layout file contains all possible rows (once we run calculate_subtotals on it), we can remove all rows where the values for a row are all zero in the results file, since when we merge we will still be able to keep those rows from the layout file in the final df. This wil help to significantly reduce the size of resutls files as well as ignore any issues that arent issues because they are all zeros. For example we had an issue where a subfuel was being sued for a subtotal where the sectors it was subtotaling didnt have that subfuel. But since the values were all zeros, it didnt matter.
    results_df = results_df.loc[~(results_df[years_to_keep_in_results] == 0).all(axis=1)].copy()
    return results_df

def check_for_differeces_between_layout_and_results_df(layout_df, filtered_results_df, shared_categories, file, CHECK_FOR_MISSING_DATA_IN_RESULTS_FILE=False):
    # Check if there are differences between the layout DataFrame and the results DataFrame. 
    differences = []
    # Compare the shared categories between layout and results DataFrame
    layout_shared_categories = layout_df[shared_categories]
    results_shared_categories = filtered_results_df[shared_categories]
    #It will check first for any categories in the results data that arent in the layout. 
    for category in shared_categories:
        diff_variables = results_shared_categories.loc[~results_shared_categories[category].isin(layout_shared_categories[category]), category].unique()
        for variable in diff_variables:
            differences.append((variable, category, 'results'))
    #then check for any categories in the layout data that arent in the results.This is pretty extreme but one day it could be useful for identifying where we could add more data to the models.
    if CHECK_FOR_MISSING_DATA_IN_RESULTS_FILE:
        for category in shared_categories:
            diff_variables = layout_shared_categories.loc[~layout_shared_categories[category].isin(results_shared_categories[category]), category].unique()
            for variable in diff_variables:
                differences.append((variable, category, 'layout'))
    # Extract the file name from the file path
    file_name = os.path.basename(file)

    # Use assert to check if there are any differences
    if len(differences) > 0:
        breakpoint()
        print(f"Differences found in results file: {file_name}\n\nDifferences:\n" + '\n'.join([f"There is no '{variable}' in '{category}' in the {'results' if df == 'layout' else 'layout'}" for variable, category, df in differences]))
        #save to a csv so we can check it later. in Outlook9th_EBT\data\temp\error_checking
        pd.DataFrame(differences, columns=['variable', 'category', 'df']).to_csv(f'data/temp/error_checking/{file_name}_differences.csv', index=False)
        print(f"Differences saved to data/temp/error_checking{file_name}_differences.csv")

# def check_for_negatives_or_postives_in_wrong_sectors(filtered_results_df, file):
    
#     # Filter rows where the sector is '05_international_aviation_bunkers'
#     sectors_to_check = ['05_international_aviation_bunkers', '04_international_marine_bunkers']
#     filtered_rows = filtered_results_df['sectors'].isin(sectors_to_check)

#     # For each year from 2021 to 2070, check if the value is positive, if yes, throw an error
#     for year in range(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR+1):
#         # Check if the value is positive
#         if (filtered_results_df.loc[filtered_rows, str(year)] > 0).any():
#             breakpoint()
#             raise Exception(f"{file} has positive values for {sectors_to_check} in {year}.")
        
def set_subfuel_x_rows_to_unallocated(concatted_results_df, layout_df):
    #take in results df with subtotals labelled and then where there is an x in the subfuels that isnt expected to be there (ie.e. x is not the most disaggregated that fuel can be) then we will rename the subfuel to 'FUEL_unallocated' so that we can use it instead of conufsing it as a subtotal later on.
    #renmae any 16_others x and 15_solid_biomass x to 16_others_unallocated and 15_solid_biomass_unallocated. this is a simple way to avoid the issues we were having. 
    known_fuels_to_make_unallocated = ['16_others', '15_solid_biomass', '01_coal', '06_crude_oil_and_ngl', '07_petroleum_products', '08_gas', '12_solar']
    fuels_with_x_as_most_disaggregated = [
        '02_coal_products',
        '03_peat',
        '04_peat_products',
        '05_oil_shale_and_oil_sands',
        '09_nuclear',
        '10_hydro',
        '11_geothermal',
        '13_tide_wave_ocean',
        '14_wind',
        '17_electricity',
        '17_x_green_electricity',
        '18_heat',
        '19_total',
        '20_total_renewables',
        '21_modern_renewables'
    ]
    new_rows_df = pd.DataFrame()
    if len(concatted_results_df.loc[(concatted_results_df['fuels'].isin(known_fuels_to_make_unallocated)) & (concatted_results_df['subfuels'] == 'x') & (concatted_results_df['is_subtotal'] == False)]) > 0:
        
        for fuel in known_fuels_to_make_unallocated:
            new_rows = concatted_results_df.loc[(concatted_results_df['fuels']==fuel) & (concatted_results_df['subfuels'] == 'x') & (concatted_results_df['is_subtotal'] == False)].copy()
            new_rows['subfuels'] = fuel+'_unallocated'
            new_rows_df = pd.concat([new_rows_df, new_rows], ignore_index=True)
            
            concatted_results_df.loc[(concatted_results_df['fuels']==fuel) & (concatted_results_df['subfuels'] == 'x') & (concatted_results_df['is_subtotal'] == False), 'subfuels'] = fuel+'_unallocated'
    if len(concatted_results_df.loc[(~concatted_results_df['fuels'].isin(fuels_with_x_as_most_disaggregated)) & (concatted_results_df['subfuels'] == 'x') & (concatted_results_df['is_subtotal'] == False)]) > 0:
        breakpoint()#china unallocated test
        fuels = concatted_results_df.loc[(~concatted_results_df['fuels'].isin(fuels_with_x_as_most_disaggregated)) & (concatted_results_df['subfuels'] == 'x') & (concatted_results_df['is_subtotal'] == False), 'fuels'].unique().tolist()
        breakpoint()
        raise Exception("There are still subfuels with x in them that are not subtotals. This is unexpected. the fuels are {}".format(fuels))
    if new_rows_df.shape[0] == 0:
        return concatted_results_df, layout_df
    
    #the below shouldnt be needed if the unallocated types are added to the layout file configs from the beginning but its useful to double check. 
    
    #add tehse new unallocated rows to the layout_df as 0s
    layout_df_years =[str(year) for year in range(EBT_EARLIEST_YEAR, OUTLOOK_LAST_YEAR+1)]
    similar_cols = [col for col in layout_df.columns if col in concatted_results_df.columns]
    similar_cols_non_years = [col for col in similar_cols if col not in layout_df_years]
    missing_cols = [col for col in layout_df.columns if col not in concatted_results_df.columns]
    
    # if any of missing cols are not in layout_df_years then throw an error else, fill all year cols with 0s
    if len([col for col in missing_cols if col not in layout_df_years]) > 0:
        breakpoint()
        raise Exception("There are missing columns in the layout_df that are not in the results_df. This is unexpected.")
    
    new_rows_df[layout_df_years] = 0
    #identify rows that are not already in the layout_df
    new_rows_df = new_rows_df[layout_df.columns].copy()
    merged_df = new_rows_df.merge(layout_df[similar_cols_non_years], on=similar_cols_non_years, how='left', indicator=True, suffixes=('', '_layout'))
    
    #drop rows that are already in the layout_df and cols with _layout in them
    merged_df = merged_df[merged_df['_merge'] == 'left_only'].copy()
    merged_df.drop(columns=['_merge'], inplace=True)
    #if tehre are any cols with _layout, throw an error
    if len([col for col in merged_df.columns if '_layout' in str(col)]) > 0:
        breakpoint()
        raise Exception("There are columns with '_layout' in them. This is unexpected.")
    #concat these new rows to the layout_df
    new_layout_df = pd.concat([layout_df, merged_df[layout_df.columns]], ignore_index=True)  
    if merged_df[layout_df.columns].shape[0]== 0:
        return concatted_results_df, layout_df
    return concatted_results_df, new_layout_df
             
                 
def trim_layout_before_merging_with_results(layout_df, concatted_results_df):
    """Trim the layout DataFrame before merging with the results DataFrame to cut down on time. This especially involves removing sectors that are not in the results DataFrame. the missing sectors are returned in a DataFrame so the user can see what is missing, and they will be concatted on after the merge.

    Args:
        layout_df (_type_): _description_
        concatted_results_df (_type_): _description_

    Returns:
        _type_: _description_
    """
    # Get the unique sectors from the results_df
    sectors_list = concatted_results_df['sectors'].unique().tolist()

    # Create a new DataFrame with rows that match the sectors from the results DataFrame. This allows us to run the merge on a results DataFrame that doesnt contain all the sectors. This is expected to happen because we will run this code multiple times, with only demand sectors, demand and transformation and so on.
    new_layout_df = layout_df[layout_df['sectors'].isin(sectors_list)].copy()

    #Drop the rows that were updated in the new DataFrame from the original layout DataFrame
    missing_sectors_df = layout_df[~layout_df['sectors'].isin(sectors_list)].copy()
    #Show user what sectors are missing:
    if missing_sectors_df.shape[0] > 0:
        print("The following sectors are missing from the results files:")
        print(missing_sectors_df['sectors'].unique().tolist())
    
    return new_layout_df, missing_sectors_df

def trim_results_before_merging_with_layout(concatted_results_df,shared_categories):
    #drop years not in the outlook (they are in the layout file)
    
    
    years_to_keep = [str(year) for year in range(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR+1)]
    # concatted_results_df = concatted_results_df[shared_categories + ['is_subtotal'] + years_to_keep].copy()#dont think this is necessary
    # Drop rows with NA or zeros in the year columns from concatted_results_df
    concatted_results_df.dropna(subset=years_to_keep, how='all', inplace=True)
    concatted_results_df = concatted_results_df.loc[~(concatted_results_df[years_to_keep] == 0).all(axis=1)]

    # Check for duplicate rows in concatted_results_df based on shared_categories
    duplicates = concatted_results_df[concatted_results_df.duplicated(subset=shared_categories, keep=False)]
    # if there are duplcaites we should let the user know.
    if duplicates.shape[0] > 0:
        print("Duplicates found in concatted_results_df based on shared_categories:")
        print(duplicates)
        
    # Remove the duplicate rows from concatted_results_df
    concatted_results_df = concatted_results_df.drop_duplicates(subset=shared_categories, keep='first').copy()

    # Print the updated number of rows in concatted_results_df
    print("Number of rows in concatted_results_df after removing rows without year values and duplicates:", concatted_results_df.shape[0])
    
    return concatted_results_df

def format_merged_layout_results_df(merged_df, shared_categories, trimmed_layout_df, trimmed_concatted_results_df, missing_sectors_df):
    #we may have found that during the merge, some rows  subtotals and some arent
    merged_df['subtotal_layout'] = merged_df['subtotal_layout'].fillna(False)
    merged_df['subtotal_results'] = merged_df['subtotal_results'].fillna(False)
    missing_sectors_df['subtotal_results'] = False
    
    #Check for duplicate rows in concatted_results_df. if there are something is wrong as
    duplicates = merged_df[merged_df.duplicated(subset=shared_categories, keep=False)]
    if duplicates.shape[0] > 0:
        breakpoint()
        raise Exception("Duplicate rows found in concatted_results_df. Check the results files.")

    # Check if there were any unexpected or extra rows in concatted_results_df
    unexpected_rows = merged_df[merged_df['_merge'] != 'both']
    if unexpected_rows.shape[0] > 0:
        missing_from_results_df = unexpected_rows[unexpected_rows['_merge'] == 'left_only']
        #find total value of data in OUTLOOK_BASE_YEAR. this shows how much worth of data might be missing from the results file.
        missing_from_results_df_value = missing_from_results_df[str(OUTLOOK_BASE_YEAR)].sum()
        extra_in_results_df = unexpected_rows[unexpected_rows['_merge'] == 'right_only']
        #find total value of data in OUTLOOK_BASE_YEAR+1
        extra_in_results_df_value = extra_in_results_df[str(OUTLOOK_BASE_YEAR+1)].sum()
        
        print(f"Unexpected rows found in concatted_results_df. Check the results files.\nMissing from results_df: {missing_from_results_df.shape[0]}, with value {missing_from_results_df_value}\nExtra in results_df: {extra_in_results_df.shape[0]}, with value {extra_in_results_df_value}")#TODO IS THIS GOING TO BE TOO STRICT? WILL IT INCLUDE FUEL TYPES THAT SHOULD BE MISSING EG. CRUDE IN TRANSPORT, or even where there is no subttotals to create in the layout file since it is not specific enough (eg breaking domestic air into freight/passenger in transport sector)
        # breakpoint()
        # raise Exception("Unexpected rows found in concatted_results_df. Check the results files.")
    
    # Print the number of rows in both dataframes
    print("Number of rows in new_layout_df:", trimmed_layout_df.shape[0])
    print("Number of rows in merged_results_df:", trimmed_concatted_results_df.shape[0])
    print("Number of rows in merged_df:", merged_df.shape[0])

    #drop the _merge column
    merged_df.drop(columns=['_merge'], inplace=True)
    
    # Combine the remainign rows form the original layout_df with the merged_df
    results_layout_df = pd.concat([missing_sectors_df, merged_df])
    return results_layout_df

def calculate_sector_aggregates(df, sectors, aggregate_sector, shared_categories, shared_categories_w_subtotals,MAJOR_SUPPLY_DATA_AVAILABLE_OVERRIDE=False):
    """
    Calculate common aggregates for groups of sectors such as total transformation,
    total primary energy supply, etc. This is based on filtering out subtotal entries 
    and grouping by scenarios, economy, fuels, and subfuels.

    Args:
        df (pd.DataFrame): The dataframe containing energy data.
        sectors (list): The sectors to include in the aggregation.
        aggregate_sector (str): The specific aggregate to calculate.

    Returns:
        pd.DataFrame: The aggregated data for the desired sectors.
    """
    # Create lists of year columns for each condition using floats
    layout_years = [str(year) for year in range(EBT_EARLIEST_YEAR, OUTLOOK_BASE_YEAR + 1)]
    results_years = [str(year) for year in range(OUTLOOK_BASE_YEAR + 1, OUTLOOK_LAST_YEAR + 1)]

    # Split the DataFrame into two based on 'subtotal_layout' and 'subtotal_results'
    df_filtered_layout = df[(df['subtotal_layout'] == False) & (df['sectors'].isin(sectors))].copy()
    df_filtered_results = df[(df['subtotal_results'] == False) & (df['sectors'].isin(sectors))].copy()

    # Drop the years not relevant for each DataFrame
    df_filtered_layout = df_filtered_layout[shared_categories_w_subtotals + layout_years]
    df_filtered_results = df_filtered_results[shared_categories_w_subtotals + results_years]

    # Define key columns for grouping
    key_cols = ['scenarios', 'economy', 'fuels', 'subfuels']

    # Initialize an empty DataFrame to store aggregated results
    aggregated_dfs = []

    # Iterate over the two filtered DataFrames
    for df_filtered in [df_filtered_layout, df_filtered_results]:
        # Assigning names to the DataFrames for identification
        df_filtered.name = 'layout' if df_filtered is df_filtered_layout else 'results'

        # If calculating total primary energy supply, make all TFC components negative first, then add them to the other components.
        if aggregate_sector == '07_total_primary_energy_supply':
            #TPES needs to be calauclate as ??? (TODO write the process for calcualting tpes so the context for below is explained)
            
            # Calculate TPES bottom up
            # Filter df to only include the demand sectors
            demand_sectors_df = df_filtered[df_filtered['sectors'].isin(['14_industry_sector', '15_transport_sector', '16_other_sector', '17_nonenergy_use'])].copy()
            tfc_df = demand_sectors_df.groupby(key_cols).sum(numeric_only=True).reset_index()
            # Multiplies all numeric columns by -1 to turn the values into negative
            numeric_cols = tfc_df.select_dtypes(include=[np.number]).columns
            
            tfc_df[numeric_cols] = tfc_df[numeric_cols] * -1
            # Filter df to only include the transformation sector
            transformation_sector_df = df_filtered[df_filtered['sectors'].isin(['08_transfers', '09_total_transformation_sector', '10_losses_and_own_use', '11_statistical_discrepancy'])].copy()
            # Concatenating the two DataFrames
            tpes_df = pd.concat([transformation_sector_df, tfc_df], ignore_index=True)
        
            tpes_bottom_up_df = tpes_df.groupby(key_cols).sum(numeric_only=True).reset_index()
            # Multiplying all numeric columns by -1 to flip the values
            numeric_cols = tpes_bottom_up_df.select_dtypes(include=[np.number]).columns
            
            tpes_bottom_up_df[numeric_cols] = tpes_bottom_up_df[numeric_cols] * -1
            
            # tpes_bottom_up_df.to_csv('data/temp/error_checking/tpes_bottom_up_' + df_filtered.name + '.csv', index=False)
            
            # Calculate TPES top down
            # Filter df to only include the supply sectors
            supply_sectors_df = df_filtered[df_filtered['sectors'].isin(['01_production','02_imports', '03_exports', '06_stock_changes', '04_international_marine_bunkers','05_international_aviation_bunkers'])].copy()
            # # Define TFC sectors and negate their values
            # negative_sectors = ['03_exports', '04_international_marine_bunkers','05_international_aviation_bunkers' ]
            # # Make the values in all numeric columns where the sector is a negative sector, negative (Don't need to do this as these values are already negative, so just multiply by 1 to keep them negative or take it out in the future)
            # negative_df = supply_sectors_df[supply_sectors_df['sectors'].isin(negative_sectors)].copy()
            # numeric_cols = negative_df.select_dtypes(include=[np.number]).columns
            # negative_df[numeric_cols] *= 1
            # supply_df = pd.concat([supply_sectors_df[~supply_sectors_df['sectors'].isin(negative_sectors)], negative_df], ignore_index=True).copy()
            
            # tpes_top_down_df = supply_df.groupby(key_cols).sum(numeric_only=True).reset_index()
            
            
            tpes_top_down_df = supply_sectors_df.groupby(key_cols).sum(numeric_only=True).reset_index().copy()
            
            # Identify numeric columns for comparison
            numeric_cols = tpes_top_down_df.select_dtypes(include=[np.number]).columns

            # Initialize a dataframe to store differences
            differences = pd.DataFrame()

            # Iterate over each row in the first dataframe
            for idx, row in tpes_top_down_df.iterrows():
                # Find the corresponding row in the second dataframe
                corresponding_row = tpes_bottom_up_df[(tpes_bottom_up_df[key_cols] == row[key_cols]).all(axis=1)]

                if not corresponding_row.empty:
                    # Compare numeric columns within tolerance
                    for col in numeric_cols:
                        value1 = row[col]
                        value2 = corresponding_row[col].values[0]
                        # Check if both values are not NaN or zero, and if difference exceeds tolerance
                        if not (np.isnan(value1) or np.isnan(value2) or value1 == 0 or value2 == 0):
                            if np.abs(value1 - value2) > 100000:
                                # Save the differing rows
                                differences = pd.concat([differences, pd.DataFrame([row]), corresponding_row])

            # Remove duplicates if any
            differences = differences.drop_duplicates()

            # Check and save the differences
            if not differences.empty:
                differences.to_csv('data/temp/error_checking/tpes_differences_' + df_filtered.name + '.csv', index=False)
                breakpoint()
                # raise Exception(f"Differences found in TPES calculation with {df_filtered.name} DataFrame and saved to 'tpes_differences_{df_filtered.name}.csv'.")
            else:
                print(f"No significant differences found in TPES calculation with {df_filtered.name} DataFrame.")
                # If MAJOR_SUPPLY_DATA_AVAILABLE is True, use the top-down TPES calculation
                if MAJOR_SUPPLY_DATA_AVAILABLE or MAJOR_SUPPLY_DATA_AVAILABLE_OVERRIDE:
                    aggregated_df = tpes_top_down_df.copy()
                else:
                    aggregated_df = tpes_bottom_up_df.copy()
                aggregated_df['subtotal_layout'] = False
                aggregated_df['subtotal_results'] = False
        #THIS COMMENTED SECTION WAS MOVEDTO create_renewables_production_for_power_transformation_input()
        # elif aggregate_sector == '01_production':
        #     if df_filtered.name == 'layout':
        #         # Filter '01_production' for df_filtered_layout
        #         aggregated_df = df_filtered[df_filtered['sectors'] == '01_production'].copy()
        #     elif df_filtered.name == 'results':
        #         # Filter '09_total_transformation_sector' for df_filtered_results
        #         df_transformation = df_filtered[df_filtered['sectors'] == '09_total_transformation_sector'].copy()
        #         df_filtered = df_filtered[df_filtered['sectors'] == '01_production'].copy()
        #         # Change all values into positive
        #         numeric_cols = df_transformation.select_dtypes(include=[np.number]).columns
        #         df_transformation[numeric_cols] = df_transformation[numeric_cols].abs()
        #         # Filter for transformation inputs that dont already have their supply modelled:
        #         df_transformation = df_transformation[df_transformation['fuels'].isin(['09_nuclear', '10_hydro', '11_geothermal', '12_solar', '13_tide_wave_ocean', '14_wind']) | df_transformation['subfuels'].isin(['16_02_industrial_waste', '16_03_municipal_solid_waste_renewable', '16_04_municipal_solid_waste_nonrenewable', '16_08_other_liquid_biofuels', '16_09_other_sources'])].copy()
        #         # Concatenate the two DataFrames
        #         df_filtered = pd.concat([df_filtered, df_transformation], ignore_index=True)
        #         # Group by key columns and sum the values
        #         aggregated_df = df_filtered.groupby(key_cols).sum(numeric_only=True).reset_index().copy()
            
        elif aggregate_sector in ['13_total_final_energy_consumption', '12_total_final_consumption']:#these also need to ahve values calcualted for fuel subtotals, like TPES
            # If not calculating total primary energy supply, just perform the grouping and sum
            aggregated_df = df_filtered.groupby(key_cols).sum(numeric_only=True).reset_index().copy()
            
        else:
            raise Exception(f'Aggregate sector {aggregate_sector} not recognised')
        
        # Add missing columns with 'x'
        for col in shared_categories:
            if col not in aggregated_df.columns.to_list() + ['subtotal_layout', 'subtotal_results']:
                with warnings.catch_warnings():
                    warnings.filterwarnings("ignore", message="DataFrame is highly fragmented")
                    aggregated_df[col] = 'x'
        
        # Reorder columns to match shared_categories order
        ordered_columns = shared_categories + [col for col in aggregated_df.columns if col not in shared_categories]
        aggregated_df = aggregated_df[ordered_columns]

        # Update the 'sectors' column
        aggregated_df['sectors'] = aggregate_sector
            
        #check fr duplicates
        dupes = aggregated_df[aggregated_df.duplicated(subset=shared_categories, keep=False)]
        if dupes.shape[0] > 0:
            print(dupes)
            breakpoint()
            raise Exception("Duplicates found in aggregated_df. Check the results files.")

        #where subtotal_layout or subtotal_results is 1 or 0 replace with True or False #TODO CLEAN THIS UP. SHOULDNT HAPPEN IN FIRST PLACE
        aggregated_df['subtotal_layout'] = aggregated_df['subtotal_layout'].replace({0:False, 1:True})
        aggregated_df['subtotal_results'] = aggregated_df['subtotal_results'].replace({0:False, 1:True})
        
        aggregated_dfs.append(aggregated_df)
    
    # Concatenate the processed DataFrames
    final_aggregated_df = pd.merge(aggregated_dfs[0], aggregated_dfs[1], on=shared_categories_w_subtotals, how='outer')
    
    #make the year cols into strs
    final_aggregated_df.columns = final_aggregated_df.columns.astype(str)
    
    return final_aggregated_df

def calculating_subtotals_in_sector_aggregates(df, shared_categories_w_subtotals):
    excluded_cols = ['subfuels']
    
    # Generate year columns as floats
    year_columns = [str(year) for year in range(EBT_EARLIEST_YEAR, OUTLOOK_LAST_YEAR + 1)]
    
    # Remove 'subfuels' from the shared categories list
    group_columns = [cat for cat in shared_categories_w_subtotals if cat not in excluded_cols]
    
    # Reset index if the DataFrame has an unnamed index column
    if df.index.name is None and not isinstance(df.index, pd.RangeIndex):
        df = df.reset_index(drop=True)
    
    # Check for missing columns
    missing_columns = [col for col in group_columns + year_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Not all specified columns are in the DataFrame. Missing columns: {missing_columns}")

    # Aggregate the data
    aggregated_df = df.groupby(group_columns).sum(numeric_only=True).reset_index().copy()

    # Add missing columns with 'x'
    for col in shared_categories_w_subtotals:
        if col not in aggregated_df.columns.to_list() + ['subtotal_layout', 'subtotal_results']:
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", message="DataFrame is highly fragmented")
                aggregated_df[col] = 'x'

    # Reorder columns to match shared_categories order
    ordered_columns = shared_categories_w_subtotals + [col for col in aggregated_df.columns if col not in shared_categories_w_subtotals]
    aggregated_df = aggregated_df[ordered_columns]

    # Drop rows in df where 'subfuels' is exactly 'x'
    df = df[df['subfuels'] != 'x']

    # Concatenate df and aggregated_df
    final_df = pd.concat([df, aggregated_df])
    
    # Drop the 'subtotal_layout' and 'subtotal_results' columns to relabel them
    final_df = final_df.drop(['subtotal_layout', 'subtotal_results'], axis=1).copy()

    return final_df

def create_final_energy_df(sector_aggregates_df, fuel_aggregates_df,results_layout_df, shared_categories):
    final_df = pd.concat([sector_aggregates_df, fuel_aggregates_df, results_layout_df], ignore_index=True)
    try:
        #replace any nas in years cols with 0
        final_df.loc[:,[str(year) for year in range(EBT_EARLIEST_YEAR, OUTLOOK_LAST_YEAR+1)]] = final_df.loc[:,[str(year) for year in range(EBT_EARLIEST_YEAR, OUTLOOK_LAST_YEAR+1)]].fillna(0)
    except:
        breakpoint()
        final_df.loc[:,[str(year) for year in range(EBT_EARLIEST_YEAR, OUTLOOK_LAST_YEAR+1)]] = final_df.loc[:,[str(year) for year in range(EBT_EARLIEST_YEAR, OUTLOOK_LAST_YEAR+1)]].fillna(0)
    #ceck for duplicates
    final_df_dupes = final_df[final_df.duplicated(subset=shared_categories, keep=False)]
    if final_df_dupes.shape[0] > 0:
        print(final_df_dupes)
        breakpoint()
        raise Exception("Duplicates found in final_df. Check the results files.")
    
    return final_df
    
    
def calculate_fuel_aggregates(new_aggregates_df, results_layout_df, shared_categories):
    """19_total is used to sum up all energy for all fuels in any sector. THis is a useful aggregate to basically check the sum of all eneryg in each sector/subsector
    
    21_modern_renewables is used to sum up all modern renewables in any sector. This is a useful aggregate to check the sum of all modern renewables in each sector/subsector. Modern renewables are any renewables that arent biomass consumed in the buildings and ag sectors.
    
    20_total_renewables is used to sum up all renewables in any sector. This is a useful aggregate to check the sum of all renewables in each sector/subsector

    Args:
        df (_type_): _description_
        shared_categories (_type_): _description_

    Raises:
        Exception: _description_

    Returns:
        _type_: _description_
    """
    #TODO ADD IN TOTAL RENEWABLES AND MODERN RENEWABLES. THEY FOLLOW SAME PATTERN (I.E CALCUALTED FOR EVERY GROUP EXCEPT FUELS, AND THEN ADDED TO THE DF)

    # # Concatenate new_aggregates_df and the original df (with no aggregates)
    df = pd.concat([new_aggregates_df, results_layout_df], ignore_index=True)
    
    # Melt the DataFrame
    df_melted = df.melt(id_vars=shared_categories + ['subtotal_layout', 'subtotal_results'], value_vars=[col for col in df.columns if col in [str(year) for year in range(EBT_EARLIEST_YEAR, OUTLOOK_LAST_YEAR+1)]],var_name='year', value_name='value')

    # Convert year column to integer
    df_melted['year'] = df_melted['year'].astype(int)

    # drop 08_02_interproduct_transfers, 09_12_nonspecified_transformation as we dont want to include it in the aggregates
    df_melted = df_melted[~df_melted['sub1sectors'].isin(['09_12_nonspecified_transformation', '08_02_interproduct_transfers'])].copy()#NOTE IM NOT REALLY SURE WHY THIS IS DONE. SEEMS VERY SPECIFIC AND BEYOND THE SCOPE OF THIS SYSTEM?

    # Split the melted DataFrame into two based on the year ranges
    df_layout = df_melted[(df_melted['year'].between(EBT_EARLIEST_YEAR, OUTLOOK_BASE_YEAR)) & (df_melted['subtotal_layout'] == False)].copy()
    df_results = df_melted[(df_melted['year'].between(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR)) & (df_melted['subtotal_results'] == False)].copy()

    def process_df_19_total(split_df):
        fuel_aggregates_list = []
        for cols_to_exclude in exclusion_sets:
            excluded_cols = ['fuels', 'subfuels'] + cols_to_exclude
            group_columns = [cat for cat in shared_categories if cat not in excluded_cols] + ['year']

            # Aggregate '19_total'
            total_19 = split_df.groupby(group_columns)['value'].sum().reset_index()
            total_19['fuels'] = '19_total'
            for col in excluded_cols[1:]:
                total_19[col] = 'x'

            fuel_aggregates_list.append(total_19)

        fuel_aggregates = pd.concat(fuel_aggregates_list, ignore_index=True)
        
        # Remove exact duplicates, keeping the first occurrence
        fuel_aggregates = fuel_aggregates.drop_duplicates()

        # Remove rows where 'value' is 0
        fuel_aggregates = fuel_aggregates[fuel_aggregates['value'] != 0]
        
        return fuel_aggregates
    
    def process_df_20_total_renewables(split_df):
        total_renewables_df = split_df[split_df['fuels'].isin(['10_hydro', '11_geothermal', '12_solar', '13_tide_wave_ocean', '14_wind', '15_solid_biomass', '16_others'])].copy()
        total_renewables_df = total_renewables_df[~total_renewables_df['subfuels'].isin(['16_02_industrial_waste', '16_04_municipal_solid_waste_nonrenewable', '16_09_other_sources', '16_x_hydrogen', '16_x_ammonia'])].copy()
        
        fuel_aggregates_list = []
        for cols_to_exclude in exclusion_sets:
            excluded_cols = ['fuels', 'subfuels'] + cols_to_exclude
            group_columns = [cat for cat in shared_categories if cat not in excluded_cols] + ['year']

            # Aggregate '20_total_renewables'
            renewables_20 = total_renewables_df.groupby(group_columns)['value'].sum().reset_index()
            renewables_20['fuels'] = '20_total_renewables'
            for col in excluded_cols[1:]:
                renewables_20[col] = 'x'

            fuel_aggregates_list.append(renewables_20)

        fuel_aggregates = pd.concat(fuel_aggregates_list, ignore_index=True)
        
        # Remove exact duplicates, keeping the first occurrence
        fuel_aggregates = fuel_aggregates.drop_duplicates()

        # Remove rows where 'value' is 0
        fuel_aggregates = fuel_aggregates[fuel_aggregates['value'] != 0]
        
        return fuel_aggregates
    
    def process_df_21_modern_renewables(split_df, df_19_total, df_20_total_renewables):
        # Filter out '18_electricity_output_in_gwh' and '19_heat_output_in_pj' from df_19_total and df_20_total_renewables for renewable shares calculation
        df_19_total_filtered = df_19_total[(df_19_total['sectors'] == '18_electricity_output_in_gwh') & (df_19_total['sub1sectors'] == 'x') | (df_19_total['sectors'] == '19_heat_output_in_pj') & (df_19_total['sub1sectors'] == 'x')].copy()
        df_20_total_renewables_filtered = df_20_total_renewables[(df_20_total_renewables['sectors'] == '18_electricity_output_in_gwh') & (df_20_total_renewables['sub1sectors'] == 'x') | (df_20_total_renewables['sectors'] == '19_heat_output_in_pj') & (df_20_total_renewables['sub1sectors'] == 'x')].copy()

        # Filter out split_df to include all possible renewables
        modern_renewables_df = split_df[split_df['fuels'].isin(['10_hydro', '11_geothermal', '12_solar', '13_tide_wave_ocean', '14_wind', '15_solid_biomass', '16_others', '17_electricity', '18_heat'])].copy()
        modern_renewables_df = modern_renewables_df[~modern_renewables_df['subfuels'].isin(['16_02_industrial_waste', '16_04_municipal_solid_waste_nonrenewable', '16_09_other_sources', '16_x_hydrogen', '16_x_ammonia'])].copy()
        # Filter out biomass consumed in buildings and agriculture
        modern_renewables_df = modern_renewables_df[~((modern_renewables_df['fuels'] == '15_solid_biomass') & (modern_renewables_df['sectors'] == '16_other_sector'))].copy()
        
        # Filter out some subsectors
        modern_renewables_df = modern_renewables_df[~((modern_renewables_df['sub1sectors'] == '09_04_electric_boilers') | (modern_renewables_df['sub2sectors'] == '10_01_02_gas_works_plants'))].copy()
        
        # Filter out '12_total_final_consumption' and '13_total_final_energy_consumption' to exclude biomass consumed in buildings and agriculture
        # modern_renewables_df = modern_renewables_df[~((modern_renewables_df['sectors'] == '12_total_final_consumption') | (modern_renewables_df['sectors'] == '13_total_final_energy_consumption'))].copy()

        # Calculate the shares
        shares_df = df_20_total_renewables_filtered[['year', 'scenarios', 'sectors', 'value']].merge(df_19_total_filtered[['year', 'scenarios', 'sectors', 'value']], on=['year', 'scenarios', 'sectors'], suffixes=('_20', '_19'))
        shares_df['share'] = shares_df['value_20'] / shares_df['value_19']
        # Add a 'fuels' column based on 'sectors' to tell if the share is for electricity or heat and be able to adjust electricity and heat values easily
        shares_df['fuels'] = shares_df['sectors'].apply(lambda x: '17_electricity' if x == '18_electricity_output_in_gwh' else ('18_heat' if x == '19_heat_output_in_pj' else None))
        shares_df.drop(columns=['sectors'], inplace=True)
        
        # Merge the share into modern_renewables_df for adjustment
        modern_renewables_df = modern_renewables_df.merge(shares_df[['year', 'scenarios', 'fuels', 'share']], on=['year', 'scenarios', 'fuels'], how='left')
        
        # Adjust values for '17_electricity' and '18_heat'
        modern_renewables_df.loc[modern_renewables_df['fuels'].isin(['17_electricity', '18_heat']), 'value'] *= modern_renewables_df['share']
        
        # Drop the 'share' column before continuing to the aggregation part
        modern_renewables_df.drop(columns=['share'], inplace=True)

        fuel_aggregates_list = []
        for cols_to_exclude in exclusion_sets:
            excluded_cols = ['fuels', 'subfuels'] + cols_to_exclude
            group_columns = [cat for cat in shared_categories if cat not in excluded_cols] + ['year']

            # Aggregate '21_modern_renewables'
            renewables_21 = modern_renewables_df.groupby(group_columns)['value'].sum().reset_index()
            renewables_21['fuels'] = '21_modern_renewables'
            for col in excluded_cols[1:]:
                renewables_21[col] = 'x'

            fuel_aggregates_list.append(renewables_21)

        fuel_aggregates = pd.concat(fuel_aggregates_list, ignore_index=True)
        
        # Remove exact duplicates, keeping the first occurrence
        fuel_aggregates = fuel_aggregates.drop_duplicates()

        # Remove rows where 'value' is 0
        fuel_aggregates = fuel_aggregates[fuel_aggregates['value'] != 0]
        
        return fuel_aggregates
    
    # Define sets of columns to exclude in each iteration
    exclusion_sets = [
        ['sub1sectors', 'sub2sectors', 'sub3sectors', 'sub4sectors'],
        ['sub2sectors', 'sub3sectors', 'sub4sectors'],
        ['sub3sectors', 'sub4sectors'],
        ['sub4sectors'],
        []
    ]
    
    def replace_values_of_tfc_and_tfec(input_df):
        # This function replaces the tfc and tfec values with the sum of the modern renewables of the sectors to exclude biomass in buildings and agriculture in modern renewables

        # Identify the rows for summing up the 'value'
        sum_df = input_df[input_df['sectors'].isin(['14_industry_sector', '15_transport_sector', '16_other_sector'])]
        sum_df = sum_df.groupby(['year', 'scenarios', 'sub1sectors', 'sub2sectors', 'sub3sectors', 'sub4sectors'], as_index=False)['value'].sum().copy()

        # Define a function to apply the sum to the 'value' column where sectors are '12_total_final_consumption' or '13_total_final_energy_consumption'
        def apply_sum(row, sum_df):
            if row['sectors'] in ['12_total_final_consumption', '13_total_final_energy_consumption']:
                sum_row = sum_df[(sum_df['year'] == row['year']) & (sum_df['scenarios'] == row['scenarios']) & (sum_df['sub1sectors'] == row['sub1sectors']) & (sum_df['sub2sectors'] == row['sub2sectors']) & (sum_df['sub3sectors'] == row['sub3sectors']) & (sum_df['sub4sectors'] == row['sub4sectors'])]
                if not sum_row.empty:
                    return sum_row['value'].values[0]
            return row['value']
        
        # Apply the function to the 'value' column
        input_df['value'] = input_df.apply(apply_sum, axis=1, sum_df=sum_df)

        return input_df
    
    # Process each split DataFrame and concatenate
    fuel_aggregates_layout_19 = process_df_19_total(df_layout)
    fuel_aggregates_results_19 = process_df_19_total(df_results)
    fuel_aggregates_layout_20 = process_df_20_total_renewables(df_layout)
    fuel_aggregates_results_20 = process_df_20_total_renewables(df_results)
    
    fuel_aggregates_layout_21 = process_df_21_modern_renewables(df_layout, fuel_aggregates_layout_19, fuel_aggregates_layout_20)
    fuel_aggregates_layout_21_modified = replace_values_of_tfc_and_tfec(fuel_aggregates_layout_21)
    fuel_aggregates_results_21 = process_df_21_modern_renewables(df_results, fuel_aggregates_results_19, fuel_aggregates_results_20)
    fuel_aggregates_results_21_modified = replace_values_of_tfc_and_tfec(fuel_aggregates_results_21)
    
    fuel_aggregates_df = pd.concat([fuel_aggregates_layout_19, fuel_aggregates_results_19,
                                    fuel_aggregates_layout_20, fuel_aggregates_results_20,
                                    fuel_aggregates_layout_21_modified, fuel_aggregates_results_21_modified], ignore_index=True)

    #######################################################
    # Temp fix for 19_total, 20_total_renewables and 21_modern_renewables in 01_production in projected years
    fuel_aggregates_df = fuel_aggregates_df[~((fuel_aggregates_df['year'].between(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR)) & (fuel_aggregates_df['sectors'] == '01_production') & (fuel_aggregates_df['fuels'].isin(['19_total', '20_total_renewables', '21_modern_renewables'])))].copy()
    # Filter for 01_production and subfuels 'x'
    df_production = df_melted[(df_melted['year'].between(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR)) & (df_melted['sectors'] == '01_production') & (df_melted['subfuels'] == 'x')].copy()
    # Sum the values for '19_total'
    total_19_production = df_production.groupby([x for x in shared_categories + ['year'] if x != 'fuels'], as_index=False)['value'].sum().copy()
    # Add 'fuels' column
    total_19_production['fuels'] = '19_total'
    # Filter for renewables
    renewables_df_production = df_melted[(df_melted['year'].between(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR)) & (df_melted['sectors'] == '01_production') & df_melted['fuels'].isin(['10_hydro', '11_geothermal', '12_solar', '13_tide_wave_ocean', '14_wind', '15_solid_biomass', '16_others']) & (df_melted['subfuels'].isin(['16_01_biogas', '16_03_municipal_solid_waste_renewable', '16_05_biogasoline', '16_06_biodiesel', '16_07_bio_jet_kerosene', '16_08_other_liquid_biofuels', 'x']))].copy()
    # Drop if 'fuel' is '16_others' and 'subfuels' is 'x'
    renewables_df_production = renewables_df_production[~((renewables_df_production['fuels'] == '16_others') & (renewables_df_production['subfuels'] == 'x'))].copy()
    # renewables_df = df_production[df_production['fuels'].isin(['10_hydro', '11_geothermal', '12_solar', '13_tide_wave_ocean', '14_wind', '15_solid_biomass'])].copy()
    # Sum the values for '20_total_renewables' and '21_modern_renewables'
    renewables_production = renewables_df_production.groupby([x for x in shared_categories + ['year'] if x not in ['fuels', 'subfuels']], as_index=False)['value'].sum().copy()
    # Add 'fuels' column and 'subfuels' column
    renewables_production['fuels'] = '20_total_renewables'
    renewables_production['subfuels'] = 'x'
    # Copy the DataFrame for '21_modern_renewables'
    modern_renewables_production = renewables_production.copy()
    modern_renewables_production['fuels'] = '21_modern_renewables'
    # Concatenate the DataFrames back to fuel_aggregates_df
    fuel_aggregates_df = pd.concat([fuel_aggregates_df, total_19_production, renewables_production, modern_renewables_production], ignore_index=True).copy()
    
    # Temp fix for 19_total, 20_total_renewables and 21_modern_renewables in 07_total_primary_energy_supply in projected years
    fuel_aggregates_df = fuel_aggregates_df[~((fuel_aggregates_df['year'].between(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR)) & (fuel_aggregates_df['sectors'] == '07_total_primary_energy_supply') & (fuel_aggregates_df['fuels'].isin(['19_total', '20_total_renewables', '21_modern_renewables'])))].copy()
    # Filter for 07_total_primary_energy_supply and subfuels 'x'
    df_primary_energy_supply = df_melted[(df_melted['year'].between(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR)) & (df_melted['sectors'] == '07_total_primary_energy_supply') & (df_melted['subfuels'] == 'x')].copy()
    # Sum the values for '19_total'
    total_19_primary_energy_supply = df_primary_energy_supply.groupby([x for x in shared_categories + ['year'] if x != 'fuels'], as_index=False)['value'].sum().copy()
    # Add 'fuels' column
    total_19_primary_energy_supply['fuels'] = '19_total'
    # Filter for renewables
    renewables_df_tpes = df_melted[(df_melted['year'].between(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR)) & (df_melted['sectors'] == '07_total_primary_energy_supply') & df_melted['fuels'].isin(['10_hydro', '11_geothermal', '12_solar', '13_tide_wave_ocean', '14_wind', '15_solid_biomass', '16_others']) & (df_melted['subfuels'].isin(['16_01_biogas', '16_03_municipal_solid_waste_renewable', '16_05_biogasoline', '16_06_biodiesel', '16_07_bio_jet_kerosene', '16_08_other_liquid_biofuels', 'x']))].copy()
    # Drop if 'fuel' is '16_others' and 'subfuels' is 'x'
    renewables_df_tpes = renewables_df_tpes[~((renewables_df_tpes['fuels'] == '16_others') & (renewables_df_tpes['subfuels'] == 'x'))].copy()
    # Sum the values for '20_total_renewables' and '21_modern_renewables'
    renewables_primary_energy_supply = renewables_df_tpes.groupby([x for x in shared_categories + ['year'] if x not in ['fuels', 'subfuels']], as_index=False)['value'].sum().copy()
    # Add 'fuels' column and 'subfuels' column
    renewables_primary_energy_supply['fuels'] = '20_total_renewables'
    renewables_primary_energy_supply['subfuels'] = 'x'
    # Copy the DataFrame for '21_modern_renewables'
    modern_renewables_primary_energy_supply = renewables_primary_energy_supply.copy()
    modern_renewables_primary_energy_supply['fuels'] = '21_modern_renewables'
    # Concatenate the DataFrames back to fuel_aggregates_df
    fuel_aggregates_df = pd.concat([fuel_aggregates_df, total_19_primary_energy_supply, renewables_primary_energy_supply, modern_renewables_primary_energy_supply], ignore_index=True).copy()
    #######################################################

    # Pivot the aggregated DataFrame
    fuel_aggregates_pivoted = fuel_aggregates_df.pivot_table(index=shared_categories, columns='year', values='value').reset_index()

    # # Change columns to str and reorder
    # pivoted_df.columns = pivoted_df.columns.astype(str)
    # pivoted_columns_order = shared_categories + [str(year) for year in range(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR+1)]
    # fuel_aggregates_pivoted = pivoted_df[pivoted_columns_order]

    # # Merge pivoted data back into the original layout
    # fuel_aggregates_pivoted = results_layout_df.merge(pivoted_df, on=shared_categories, how='left')
    # layout_columns_order = shared_categories + [str(year) for year in range(EBT_EARLIEST_YEAR, OUTLOOK_LAST_YEAR+1)]
    # fuel_aggregates_pivoted = fuel_aggregates_pivoted[layout_columns_order]

        
    ####
    #20_total_renewables
    # fuels:
    # '01_coal', '02_coal_products', '03_peat', '04_peat_products',
    #    '05_oil_shale_and_oil_sands', '06_crude_oil_and_ngl',
    #    '07_petroleum_products', '08_gas', '09_nuclear', '10_hydro',
    #    '11_geothermal', '12_solar', '13_tide_wave_ocean', '14_wind',
    #    '15_solid_biomass', '16_others', '17_electricity', '18_heat'
    
    #subfuels:
    #array(['01_01_coking_coal', '01_05_lignite', '01_x_thermal_coal', 'x',
    #    '06_01_crude_oil', '06_02_natural_gas_liquids',
    #    '06_x_other_hydrocarbons', '07_01_motor_gasoline',
    #    '07_02_aviation_gasoline', '07_03_naphtha', '07_06_kerosene',
    #    '07_07_gas_diesel_oil', '07_08_fuel_oil', '07_09_lpg',
    #    '07_10_refinery_gas_not_liquefied', '07_11_ethane',
    #    '07_x_jet_fuel', '07_x_other_petroleum_products',
    #    '08_01_natural_gas', '08_02_lng', '08_03_gas_works_gas',
    #    '12_01_of_which_photovoltaics', '12_x_other_solar',
    #    '15_01_fuelwood_and_woodwaste', '15_02_bagasse', '15_03_charcoal',
    #    '15_04_black_liquor', '15_05_other_biomass', '16_01_biogas',
    #    '16_02_industrial_waste', '16_03_municipal_solid_waste_renewable',
    #    '16_04_municipal_solid_waste_nonrenewable', '16_05_biogasoline',
    #    '16_06_biodiesel', '16_07_bio_jet_kerosene',
    #    '16_08_other_liquid_biofuels', '16_09_other_sources',
    #    '16_x_ammonia', '16_x_efuel', '16_x_hydrogen'], dtype=object)
    
    # #the problem with renwables calcualtion is that we have to consider the renewables used in ouput of electricity and heat. this has to be done using the share of renewables in total genreation. this is so that renewables also has a share in losses and own use.  (edito said that). I dont know how to do this yet!
    # #TODO JNSUT CALCUALTE USING SAME METHOD AS 19 TOTAL, FOR NOW
    # total_renewables_20 = df_melted.groupby(group_columns)['value'].sum().reset_index()
    # total_renewables_20['fuels'] = '20_total_renewables'
    # for col in excluded_cols[1:]:  # We start from 1 as 'fuels' is already addressed above
    #     total_renewables_20[col] = 'x'
    # ####
    # #21_modern_renewables
    # #here, we do know that it is renewables that arent biomass consumed in the buildings and ag sectors. but we still dont know how to claulte the share of renewables in total generation. so for now we will just do the same as 19 total.
    # modern_renewables_21 = df_melted.groupby(group_columns)['value'].sum().reset_index()
    # modern_renewables_21['fuels'] = '21_modern_renewables'
    # for col in excluded_cols[1:]:  # We start from 1 as 'fuels' is already addressed above
    #     modern_renewables_21[col] = 'x'
        
    # ####    
    # fuel_aggregates = pd.concat([total_19, total_renewables_20, modern_renewables_21], ignore_index=True)
    ##############   
    
    # with warnings.catch_warnings():
    #     warnings.filterwarnings("ignore", message="DataFrame is highly fragmented")
    #     Pivot the dataframe back to its original format
    #     fuel_aggregates_pivoted = fuel_aggregates.pivot(index=[col for col in fuel_aggregates.columns if col not in ['year', 'value']], columns='year', values='value').reset_index()
        
    # dupes = fuel_aggregates_pivoted[fuel_aggregates_pivoted.duplicated(subset=shared_categories, keep=False)]
    # if dupes.shape[0] > 0:
    #     print(dupes)
    #     breakpoint()
    #     raise Exception("Duplicates found in fuel_aggregates_pivoted. Check the results files.")
    
    # set subtotals to true
    fuel_aggregates_pivoted['subtotal_layout'] = False
    fuel_aggregates_pivoted['subtotal_results'] = False
    
    #make the year cols back into str
    fuel_aggregates_pivoted.columns = fuel_aggregates_pivoted.columns.astype(str)
    return fuel_aggregates_pivoted

def load_previous_merged_df(results_data_path, expected_columns, previous_merged_df_filename):
    
    if previous_merged_df_filename is None:
        #grab the most recent previous merge file
        previous_merged_df_filename = find_most_recent_file_date_id(results_data_path)
        
        if previous_merged_df_filename is None:
            print("No previous merged df found. Skipping comparison.")
            return None
        else:
            previous_merged_df = pd.read_csv(os.path.join(results_data_path, previous_merged_df_filename))
            #check the columns are as we expect:
            missing_cols = [col for col in expected_columns if col not in previous_merged_df.columns.tolist()]
            if len(missing_cols) > 0:
                print("WARNING: The previous merged df does not have the expected columns. Skipping comparison.")
                return None
            else:
                return previous_merged_df  
            
def compare_to_previous_merge(new_merged_df, shared_categories, results_data_path,previous_merged_df_filename=None, new_subtotal_columns=['subtotal_layout', 'subtotal_results'], previous_subtotal_columns=['subtotal_historic','subtotal_predicted','is_subtotal']):
    """Use this to compare to previously done merges. this is useful for checking if any changes to the code result in any changes to the ouput compared to a previous merge. it is a bit specific but possibly helpful in understanding how this complex process works!
    Args:
        new_merged_df (_type_): _description_
        shared_categories (_type_): _description_
        results_data_path (_type_): _description_
        previous_merged_df_filename (_type_, optional): _description_. Defaults to None.
        new_subtotal_columns (list, optional): _description_. Defaults to ['subtotal_layout', 'subtotal_results'].
        previous_subtotal_columns (list, optional): _description_. Defaults to ['subtotal_historic','subtotal_predicted','is_subtotal'].
    """
    #function that can be used to identify potential errors in new code or even just to check that the results are the same as before.
    if previous_merged_df_filename is None:
        expected_columns = previous_subtotal_columns + shared_categories
        previous_merged_df = load_previous_merged_df(results_data_path, expected_columns, previous_merged_df_filename)
        if previous_merged_df is None:
            return
    else:
        previous_merged_df = pd.read_csv(os.path.join(results_data_path, previous_merged_df_filename))
    
    #melt the years cols to make it easier to compare
    new_merged_df = new_merged_df.melt(id_vars=shared_categories+new_subtotal_columns, var_name='year', value_name='value')
    previous_merged_df = previous_merged_df.melt(id_vars=shared_categories + previous_subtotal_columns, var_name='year', value_name='value')
    
    #make sure both dataframes year values are ints
    new_merged_df['year'] = new_merged_df['year'].astype(int)
    previous_merged_df['year'] = previous_merged_df['year'].astype(int)
    
    #we will have to do the merge bit by bit as it is so large. so we will split it into years and then merge each year, check the data, then move on to the next year.
    left_only = pd.DataFrame()
    right_only = pd.DataFrame()
    both_different_rows = pd.DataFrame()
    all = pd.DataFrame()
    epsilon = 1e-6
    SAVE = False
    for years in np.array_split(new_merged_df['year'].unique(), 10):
        all_ = pd.merge(new_merged_df[new_merged_df['year'].isin(years)], previous_merged_df[previous_merged_df['year'].isin(years)], on=shared_categories+['year'], how='outer', indicator=True, suffixes=('_new', '_previous'))
        #check for rows that arent in both dataframes
        left_only_ = all_[all_['_merge'] == 'left_only']
        right_only_ = all_[all_['_merge'] == 'right_only']
        #where both is true, see what the difference in value sum is between the two dataframes
        all_['value_sum_difference'] = all_['value_new'] - all_['value_previous']
        #if the difference is greater than epsilon, then there is a difference between the two dataframes and separate the different values into a new dataframe
        all_['is_different'] = all_['is_different'].astype(bool)
        all_.loc[(abs(all_['value_sum_difference']) > epsilon) & (all_['_merge'] == 'both'), 'is_different'] = True
        all_['is_different'] = all_['is_different'].fillna(False)
        
        left_only = pd.concat([left_only, left_only_], ignore_index=True)
        right_only = pd.concat([right_only, right_only_], ignore_index=True)
        
        all = pd.concat([all, all_], ignore_index=True)
    
    #rename left_only and right_only to new and previous_only
    merge_dict = {'left_only': 'new_only', 'right_only': 'previous_only'}
    all['_merge'] = all['_merge'].map(merge_dict).fillna(all['_merge'])
    
    if all['is_different'].any():
        SAVE=True
        breakpoint()
        print("{} differences in values, to an absolute sum of {} between previous and new merged df saved to data/temp/error_checking/compare_to_previous_merge.csv".format(all['is_different'].sum(), all['value_sum_difference'].sum()))
    
    #check out left_only and right_only
    if left_only.shape[0] > 0:
        SAVE=True
        print("{} rows that are only in the new merged df saved in data/temp/error_checking/compare_to_previous_merge.csv where _merge is left_only".format(left_only.shape[0]))
        breakpoint()
    
    if right_only.shape[0] > 0:
        SAVE=True
        print("{} rows that are only in the previous merged df saved in data/temp/error_checking/compare_to_previous_merge.csv where _merge is right_only".format(right_only.shape[0]))
        breakpoint()
    
    if SAVE:
        all.to_csv('data/temp/error_checking/compare_to_previous_merge.csv', index=False)
    else:
        print("No differences found between previous and new merged df.")
        
    return

def check_for_issues_by_comparing_to_layout_df(results_layout_df, shared_categories, new_aggregate_sectors, layout_df_subtotals_labelled, REMOVE_LABELLED_SUBTOTALS=False):
    """Use this to check that the layout df and the newly processed layout df match for the years in the layout file. This should not happen, and if there are isues its likely some process in the merging_results script is wrong.
    The process will do two comparisons, one on the values and the next on whether any rows are missing. The one that checks values will
    
    REMOVE_LABELLED_SUBTOTALS:
    Because there are so many issues with both the input layout and input results dfs we are never going to be able to do this perfectly, but if we assume that we have labelled subttoals correctly then if we remove the subttoals from the layout df, we will have far fewer missing rows. if you are trying to track down specific bugs though, you might want to set REMOVE_LABELLED_SUBTOTALS to False"""
    #drop where there are aggregate sectors where subfuel != x as we have removed those.
    layout_df = layout_df_subtotals_labelled.copy()
    layout_df = layout_df[~((layout_df.sectors.isin(new_aggregate_sectors) & (layout_df.subfuels != 'x')))].copy()
    if REMOVE_LABELLED_SUBTOTALS:
        layout_df=layout_df[~layout_df.is_subtotal].copy()
        results_layout_df = results_layout_df[~results_layout_df.subtotal_layout].copy()
    #Check that the layout df matches aggregates_df where year is in  [year for year in range(EBT_EARLIEST_YEAR, OUTLOOK_BASE_YEAR+1) if year in layout_df.columns]
    historical_years = [str(year) for year in range(EBT_EARLIEST_YEAR, OUTLOOK_BASE_YEAR+1)]
    results_layout_df_layout_years = results_layout_df[shared_categories + [col for col in results_layout_df.columns if col in historical_years]].copy()
    # layout_df_test = layout_df[shared_categories +  [col for col in results_layout_df.columns if col in range(EBT_EARLIEST_YEAR, OUTLOOK_BASE_YEAR+1)]].copy()
    #test they match by doing an outer merge and checking where indicator != both
    shared_categories_old = shared_categories.copy()
    shared_categories_old.remove('subtotal_layout')
    shared_categories_old.remove('subtotal_results')
    merged_df_on_values = pd.merge(results_layout_df_layout_years, layout_df, on=shared_categories_old+  [col for col in results_layout_df_layout_years.columns if col in historical_years], how="outer", indicator=True)
    merged_df_on_categories_only = pd.merge(results_layout_df_layout_years, layout_df, on=shared_categories_old, how="outer", indicator=True, suffixes=('_new_layout_df', '_original_layout'))
    
    if merged_df_on_values[merged_df_on_values['_merge'] != 'both'].shape[0] > 0:
        #wrongly calculated rows occur where we have duplicates when you ignore the years. these dont show up in merged_df_on_categories_only.
        merged_df_bad_values = merged_df_on_values[merged_df_on_values['_merge'] != 'both'].copy()
        merged_df_bad_values = merged_df_bad_values[merged_df_bad_values.duplicated(subset=shared_categories_old, keep=False)].copy()
        
        #replace left_only with new_layout_df and right_only with original_layout
        merge_dict = {'left_only': 'new_layout_df', 'right_only': 'original_layout'}
        merged_df_bad_values['_merge'] = merged_df_bad_values['_merge'].map(merge_dict).fillna(merged_df_bad_values['_merge'])
        
        #idenitfy where the diffrence between values is > than epsilon. to do this we should first melt both dfs so we have a year and value col. then join on shared_categories_old and year. then we can compare the values and see where they are different.
        merged_df_bad_values = merged_df_bad_values.drop(columns=['subtotal_layout', 'subtotal_results']).copy()
        new_layout_df = merged_df_bad_values[merged_df_bad_values['_merge'] == 'new_layout_df'].copy()
        original_layout = merged_df_bad_values[merged_df_bad_values['_merge'] == 'original_layout'].copy()
        
        #melt both then join
        new_layout_df = new_layout_df.drop(columns=['_merge']).melt(id_vars=shared_categories_old, var_name='year', value_name='value')
        original_layout = original_layout.drop(columns=['_merge']).melt(id_vars=shared_categories_old, var_name='year', value_name='value')
        merged_df_bad_values = pd.merge(new_layout_df, original_layout, on=shared_categories_old+['year'], how='inner', suffixes=('_new_layout_df', '_original_layout'))
        epsilon = 1e-6
        merged_df_bad_values['difference'] = merged_df_bad_values['value_new_layout_df'] - merged_df_bad_values['value_original_layout']
        merged_df_bad_values = merged_df_bad_values[abs(merged_df_bad_values['difference']) > epsilon].copy()
        
        ###############
        bad_values_rows_exceptions_dict = {}#IGNORE ANY AGGREGATE FUELS/SECTORS UNTIL THEYVE BEEN CALCAULTED PROPERLY
        bad_values_rows_exceptions_dict['19_total'] = {'fuels':'19_total'}
        bad_values_rows_exceptions_dict['20_total_renewables'] = {'fuels':'20_total_renewables'}
        bad_values_rows_exceptions_dict['21_modern_renewables'] = {'fuels':'21_modern_renewables'}
        bad_values_rows_exceptions_dict['13_total_final_energy_consumption'] = {'sectors':'13_total_final_energy_consumption'}
        bad_values_rows_exceptions_dict['12_total_final_consumption'] = {'sectors':'12_total_final_consumption'}
        bad_values_rows_exceptions_dict['07_total_primary_energy_supply'] = {'sectors':'07_total_primary_energy_supply'}

        # THA file has some issues with the following rows
        bad_values_rows_exceptions_dict['THA_12_solar'] = {'economy':'19_THA', 'sectors':'01_production', 'sub1sectors':'x', 'fuels':'12_solar', 'subfuels':'x'}

        # USA file has some issues with the following rows
        bad_values_rows_exceptions_dict['USA_14_industry_sector'] = {'economy':'20_USA', 'sectors':'14_industry_sector', 'sub1sectors':'x', 'fuels':'07_petroleum_products', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['USA_10_losses_and_own_use'] = {'economy':'20_USA', 'sectors':'10_losses_and_own_use', 'fuels':'01_coal', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['USA_11_statistical_discrepancy'] = {'economy':'20_USA', 'sectors':'11_statistical_discrepancy', 'sub1sectors':'x', 'fuels':'16_others', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['USA_12_solar'] = {'economy':'20_USA', 'fuels':'12_solar', 'subfuels':'x'}
        
        # JPN file has some issues with the following rows
        bad_values_rows_exceptions_dict['JPN_09_total_transformation_sector'] = {'economy':'08_JPN', 'sectors':'09_total_transformation_sector', 'sub1sectors':'x', 'fuels':'12_solar', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['JPN_11_statistical_discrepancy'] = {'economy':'08_JPN', 'sectors':'11_statistical_discrepancy', 'sub1sectors':'x', 'fuels':'16_others', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['JPN_12_solar'] = {'economy':'08_JPN', 'fuels':'12_solar', 'subfuels':'x'}
        
        # CDA file has some issues with the following rows
        bad_values_rows_exceptions_dict['CDA_11_statistical_discrepancy'] = {'economy':'03_CDA', 'sectors':'11_statistical_discrepancy', 'sub1sectors':'x', 'fuels':'16_others', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['CDA_12_solar'] = {'economy':'03_CDA', 'fuels':'12_solar', 'subfuels':'x'}
        
        # ROK file has some issues with the following rows
        bad_values_rows_exceptions_dict['ROK_11_statistical_discrepancy'] = {'economy':'09_ROK', 'sectors':'11_statistical_discrepancy', 'sub1sectors':'x', 'fuels':'16_others', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['ROK_19_heat_output_in_pj'] = {'economy':'09_ROK', 'sectors':'19_heat_output_in_pj', 'sub1sectors':'x'}
        bad_values_rows_exceptions_dict['ROK_12_solar'] = {'economy':'09_ROK', 'fuels':'12_solar', 'subfuels':'x'}
        
        # AUS file has some issues with the following rows
        bad_values_rows_exceptions_dict['AUS_11_statistical_discrepancy'] = {'economy':'01_AUS', 'sectors':'11_statistical_discrepancy', 'sub1sectors':'x', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['AUS_09_total_transformation_sector'] = {'economy':'01_AUS', 'sectors':'09_total_transformation_sector', 'fuels':'12_solar', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['AUS_18_electricity_output_in_gwh'] = {'economy':'01_AUS', 'sectors':'18_electricity_output_in_gwh', 'fuels':'12_solar', 'subfuels':'x'}
        
        # SGP file has some issues with the following rows
        bad_values_rows_exceptions_dict['SGP_18_electricity_output_in_gwh'] = {'economy':'17_SGP', 'sectors':'18_electricity_output_in_gwh'}
        
        # PRC file has some issues with the following rows
        bad_values_rows_exceptions_dict['PRC_10_losses_and_own_use'] = {'economy':'05_PRC', 'sectors':'10_losses_and_own_use', 'fuels':'08_gas', 'subfuels':'x'}
        
        # NZ file has some issues with the following rows
        bad_values_rows_exceptions_dict['NZ_11_statistical_discrepancy'] = {'economy':'12_NZ', 'sectors':'11_statistical_discrepancy', 'sub1sectors':'x', 'fuels':'16_others', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['NZ_14_industry_sector'] = {'economy':'12_NZ', 'sectors':'14_industry_sector', 'sub1sectors':'14_03_manufacturing', 'sub2sectors':'x', 'fuels':'16_others', 'subfuels':'x'}
        
        # PNG file has some issues with the following rows
        bad_values_rows_exceptions_dict['PNG_09_06_gas_processing_plants'] = {'economy':'13_PNG', 'sectors':'09_total_transformation_sector', 'sub1sectors':'09_06_gas_processing_plants', 'fuels':'08_gas', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['PNG_14_industry_sector'] = {'economy':'13_PNG', 'sectors':'14_industry_sector', 'fuels':'07_petroleum_products', 'subfuels':'x'}
        
        # HKC file has some issues with the following rows
        bad_values_rows_exceptions_dict['HKC_11_statistical_discrepancy'] = {'economy':'06_HKC', 'sectors':'11_statistical_discrepancy', 'sub1sectors':'x', 'fuels':'08_gas', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['HKC_14_industry_sector'] = {'economy':'06_HKC', 'sectors':'14_industry_sector', 'fuels':'16_others', 'subfuels':'x'}
        
        # BD file has some issues with the following rows
        bad_values_rows_exceptions_dict['BD_08_transfers'] = {'economy':'02_BD', 'sectors':'08_transfers', 'sub1sectors':'x', 'fuels':'07_petroleum_products', 'subfuels':'x'}
        
        # RUS file has some issues with the following rows
        bad_values_rows_exceptions_dict['RUS_11_statistical_discrepancy'] = {'economy':'16_RUS', 'sectors':'11_statistical_discrepancy', 'sub1sectors':'x', 'subfuels':'x'}
        # bad_values_rows_exceptions_dict['RUS_19_heat_output_in_pj'] = {'economy':'16_RUS', 'sectors':'19_heat_output_in_pj', 'sub1sectors':'x'}
        
        # MEX file has some issues with the following rows
        bad_values_rows_exceptions_dict['MEX_11_statistical_discrepancy'] = {'economy':'11_MEX', 'sectors':'11_statistical_discrepancy', 'sub1sectors':'x', 'fuels':'16_others', 'subfuels':'x'}
        
        # CHL file has some issues with the following rows
        bad_values_rows_exceptions_dict['CHL_11_statistical_discrepancy'] = {'economy':'04_CHL', 'sectors':'11_statistical_discrepancy', 'sub1sectors':'x', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['CHL_12_solar'] = {'economy':'04_CHL', 'sub1sectors':'x', 'fuels':'12_solar', 'subfuels':'x'}
        
        # CT file has some issues with the following rows
        bad_values_rows_exceptions_dict['CT_01_production'] = {'economy':'18_CT', 'sectors':'01_production', 'sub1sectors':'x', 'fuels':'01_coal', 'subfuels':'x'}
        
        #######TEMP FOR NEW ESTO DATA
        #PROBALBY TEMP FOR THE PERIOD WHILE ESTO IS CLEANINGN THEIR 2022 DATA:
        # 10_MAS	10_losses_and_own_use	x	x	x	x	08_gas	x	1984-TO-2021
        # 10_MAS 11_statistical_discrepancy	x	x	x	x	08_gas	x	1984-TO-2021
        bad_values_rows_exceptions_dict['MAS_10_losses_and_own_use'] = {'economy':'10_MAS', 'sectors':'10_losses_and_own_use', 'fuels':'08_gas', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['MAS_11_statistical_discrepancy'] = {'economy':'10_MAS', 'sectors':'11_statistical_discrepancy', 'fuels':'08_gas', 'subfuels':'x'}
        
        bad_values_rows_exceptions_dict['CT_19_heat_output_in_pj'] = {'economy':'18_CT', 'sectors':'19_heat_output_in_pj', 'sub1sectors':'x', 'fuels':'16_others', 'subfuels':'x'}
        # 12_NZ	11_statistical_discrepancy	x	x	x	x	15_solid_biomass	x
        bad_values_rows_exceptions_dict['NZ_11_statistical_discrepancy_biomass'] = {'economy':'12_NZ', 'sectors':'11_statistical_discrepancy', 'fuels':'15_solid_biomass', 'subfuels':'x'}
        bad_values_rows_exceptions_dict['NZ_11_statistical_discrepancy_others'] = {'economy':'12_NZ', 'sectors':'11_statistical_discrepancy', 'fuels':'16_others', 'subfuels':'x'}
        # 19_THA	11_statistical_discrepancy	x	x	x	x	06_crude_oil_and_ngl	x
        bad_values_rows_exceptions_dict['THA_11_statistical_discrepancy'] = {'economy':'19_THA', 'sectors':'11_statistical_discrepancy', 'fuels':'06_crude_oil_and_ngl', 'subfuels':'x'}
        
        # 05_PRC	10_losses_and_own_use	x	x	x	x	01_coal	01_x_thermal_coal 1990-2022
        bad_values_rows_exceptions_dict['PRC_10_losses_and_own_use_coal'] = {'economy':'05_PRC', 'sectors':'10_losses_and_own_use', 'fuels':'01_coal'}
        # 05_PRC	10_losses_and_own_use	x	x	x	x	02_coal_products	x 1991-2022
        bad_values_rows_exceptions_dict['PRC_10_losses_and_own_use_coal_products'] = {'economy':'05_PRC', 'sectors':'10_losses_and_own_use', 'fuels':'02_coal_products'}
        #05_PRC	10_losses_and_own_use	x	x	x	x	07_petroleum_products	07_07_gas_diesel_oil and 07_09_lpg 1991-2022
        bad_values_rows_exceptions_dict['PRC_10_losses_and_own_use_petroleum_products'] = {'economy':'05_PRC', 'sectors':'10_losses_and_own_use', 'fuels':'07_petroleum_products'}
        
        # breakpoint()#consider by december 2024 whether thes are still necessary or we shoiuld fix the data. or undeerlying issue.
        #######TEMP FOR NEW ESTO DATA
        #CREATE ROWS TO IGNORE. THESE ARE ONES THAT WE KNOW CAUSE ISSUES BUT ARENT NECESSARY TO FIX, AT LEAST RIGHT NOW
        #use the keys as column names to remove the rows in the dict:
        for ignored_issue in bad_values_rows_exceptions_dict.keys():
            #iterate through the dict to thin down to the rows we want to remove and then remove them by index
            rows_to_remove = merged_df_bad_values.copy()
            for col in bad_values_rows_exceptions_dict[ignored_issue].keys():
                rows_to_remove = rows_to_remove[rows_to_remove[col] == bad_values_rows_exceptions_dict[ignored_issue][col]].copy()
            merged_df_bad_values = merged_df_bad_values.drop(rows_to_remove.index).copy()
        
        ###############
        #where we have unexpected rows, this is where either _merge is new_layout_df or original_layout in merged_df_on_categories_only. new_layout_df means the layout file has a row that isnt in the results file. original_layout means the results file has a row that isnt in the layout file.
        missing_rows = merged_df_on_categories_only[merged_df_on_categories_only['_merge'] != 'both'].copy()
        missing_rows['_merge'] = missing_rows['_merge'].map(merge_dict).fillna(missing_rows['_merge'])
        #that covers all possible issues. we will now save them and ask the user to fix them.
        if merged_df_bad_values.shape[0] > 0:
            #put them in order to help see the issue
            merged_df_bad_values.sort_values(by=shared_categories_old, inplace=True)
            economy=merged_df_bad_values['economy'].unique()[0]
            merged_df_bad_values.to_csv(f'data/temp/error_checking/merged_df_bad_values_{economy}.csv', index=False)
            print("There are {} rows where the values in the results file do not match the values in the layout file. These rows have been saved to data/temp/error_checking/merged_df_bad_values_{}.csv".format(merged_df_bad_values.shape[0], economy))
            # breakpoint()
        if missing_rows.shape[0] > 0:
            ###############
            missing_rows_exceptions_dict = {}
            #CREATE MISSING ROWS TO IGNORE. THESE ARE ONES THAT WE KNOW CAUSE ISSUES BUT ARENT NECESSARY TO FIX, AT LEAST RIGHT NOW
            missing_rows_exceptions_dict['nonspecified_transformation'] = {'_merge':'original_layout', 'sub1sectors':'09_12_nonspecified_transformation'}
            
            missing_rows_exceptions_dict['12_total_final_consumption'] = {'_merge':'new_layout_df', 'sectors':'12_total_final_consumption'}
            missing_rows_exceptions_dict['13_total_final_energy_consumption'] = {'_merge':'new_layout_df', 'sectors':'13_total_final_energy_consumption'}
            missing_rows_exceptions_dict['07_total_primary_energy_supply'] = {'sectors':'07_total_primary_energy_supply'}
            
            missing_rows_exceptions_dict['19_total'] = {'_merge':'original_layout', 'fuels':'19_total'}
            missing_rows_exceptions_dict['20_total_renewables'] = {'_merge':'original_layout', 'fuels':'20_total_renewables'}
            missing_rows_exceptions_dict['21_modern_renewables'] = {'_merge':'original_layout', 'fuels':'21_modern_renewables'}

            # USA file has some issues with the following rows (couldn't work out the cause)
            missing_rows_exceptions_dict['nonspecified_transformation3'] = {'_merge':'new_layout_df', 'economy':'20_USA', 'sub1sectors':'09_06_gas_processing_plants', 'fuels':'20_total_renewables'}
            missing_rows_exceptions_dict['nonspecified_transformation4'] = {'_merge':'new_layout_df', 'economy':'20_USA', 'sub1sectors':'09_06_gas_processing_plants', 'fuels':'21_modern_renewables'}
            missing_rows_exceptions_dict['10_losses_and_own_use'] = {'_merge':'new_layout_df', 'economy':'20_USA', 'sub1sectors':'10_01_own_use'}
            missing_rows_exceptions_dict['08_transfers'] = {'_merge':'new_layout_df', 'economy':'20_USA', 'sectors':'08_transfers'}
            # missing_rows_exceptions_dict['19_01_05_others'] = {'_merge':'new_layout_df', 'economy':'20_USA', 'sectors':'19_heat_output_in_pj', 'sub1sectors':'19_01_chp_plants', 'sub2sectors':'19_01_05_others'}

            # JPN file has some issues with the following rows
            missing_rows_exceptions_dict['19_heat_output_in_pj'] = {'_merge':'new_layout_df', 'economy':'08_JPN', 'sub1sectors':'19_02_heat_plants'}
            # missing_rows_exceptions_dict['10_losses_and_own_use'] = {'_merge':'new_layout_df', 'economy':'08_JPN', 'sectors':'10_losses_and_own_use', 'sub1sectors':'10_02_transmision_and_distribution_losses', 'sub2sectors':'x'}
            
            # ROK file has some issues with the following rows
            missing_rows_exceptions_dict['09_05_chemical_heat_for_electricity_production'] = {'_merge':'new_layout_df', 'economy':'09_ROK', 'sub1sectors':'09_05_chemical_heat_for_electricity_production'}
            
            # BD file has some issues with the following rows
            missing_rows_exceptions_dict['11_statistical_discrepancy'] = {'_merge':'new_layout_df', 'economy':'02_BD', 'sectors':'11_statistical_discrepancy', 'sub1sectors':'x', 'fuels':'21_modern_renewables', 'subfuels':'x'}
            
            # CHL file has some issues with the following rows
            missing_rows_exceptions_dict['09_total_transformation_sector'] = {'_merge':'new_layout_df', 'economy':'04_CHL', 'sectors':'09_total_transformation_sector', 'sub1sectors':'09_05_chemical_heat_for_electricity_production', 'sub2sectors':'x', 'fuels':'21_modern_renewables', 'subfuels':'x'}

            ################TEMP FOR NEW ESTO DATA
            
            missing_rows_exceptions_dict['08_transfers'] = {'_merge':'new_layout_df', 'economy':'14_PE', 'sectors':'08_transfers', 'fuels':'07_petroleum_products', 'subfuels':'07_x_other_petroleum_products', 'sub1sectors':'x', 'sub2sectors':'x', 'sub3sectors':'x', 'sub4sectors':'x'}
            # new_layout_df	reference	05_PRC	09_total_transformation_sector	09_x_heat_plants	x	x	x	15_solid_biomass	15_solid_biomass_unallocated

            missing_rows_exceptions_dict['biomass_unallocated_power']= {'_merge':'new_layout_df', 'sectors':'09_total_transformation_sector',  'fuels':'15_solid_biomass', 'subfuels':'15_solid_biomass_unallocated'}
            
            missing_rows_exceptions_dict['others_unallocated_power']= {'_merge':'new_layout_df', 'sectors':'09_total_transformation_sector',  'fuels':'16_others_unallocated', 'subfuels':'16_others_unallocated'}
            
            #also this in NEew ESTO data:
            #             09_ROK	09_total_transformation_sector	09_01_electricity_plants	09_01_11_otherfuel	x	x	16_others	16_09_other_sources
            # 09_ROK	09_total_transformation_sector	09_02_chp_plants	09_02_04_biomass	x	x	16_others	16_09_other_sources
            # # # 09_ROK	09_total_transformation_sector	09_02_chp_plants	x	x	x	16_others	16_09_other_sources
            # # missing_rows_exceptions_dict['16_09_other_sources'] = {'_merge':'new_layout_df', 'economy':'09_ROK', 'sectors':'09_total_transformation_sector', 'sub1sectors':'09_02_chp_plants', 'fuels':'16_others', 'subfuels':'16_09_other_sources'}
            # # missing_rows_exceptions_dict['16_09_other_sources2'] = {'_merge':'new_layout_df', 'economy':'09_ROK', 'sectors':'09_total_transformation_sector', 'sub2sectors':'09_01_11_otherfuel', 'fuels':'16_others', 'subfuels':'16_09_other_sources'}
            # # missing_rows_exceptions_dict['15_05_other_biomass_1'] = {'_merge':'new_layout_df', 'economy':'09_ROK', 'sectors':'09_total_transformation_sector', 'sub1sectors':'09_01_electricity_plants', 'sub2sectors':'09_01_06_biomass', 'fuels':'15_solid_biomass', 'subfuels':'15_05_other_biomass'}
            # # missing_rows_exceptions_dict['16_02_industrial_waste_1'] = {'_merge':'new_layout_df', 'economy':'09_ROK', 'sectors':'09_total_transformation_sector', 'sub1sectors':'09_01_electricity_plants', 'sub2sectors':'09_01_11_otherfuel', 'fuels':'16_others', 'subfuels':'16_02_industrial_waste'}
            # # missing_rows_exceptions_dict['15_05_other_biomass_2'] = {'_merge':'new_layout_df', 'economy':'09_ROK', 'sectors':'09_total_transformation_sector', 'sub1sectors':'09_02_chp_plants', 'sub2sectors':'09_02_04_biomass', 'fuels':'15_solid_biomass', 'subfuels':'15_05_other_biomass'}
            # # missing_rows_exceptions_dict['16_02_industrial_waste_2'] = {'_merge':'new_layout_df', 'economy':'09_ROK', 'sectors':'09_total_transformation_sector', 'sub1sectors':'09_02_chp_plants', 'sub2sectors':'09_02_04_biomass', 'fuels':'16_others', 'subfuels':'16_02_industrial_waste'}
            #these rows:
            # economy	sectors	sub1sectors	sub2sectors	sub3sectors	sub4sectors	fuels	subfuels
            # 18_CT	09_total_transformation_sector	09_06_gas_processing_plants	09_06_02_liquefaction_regasification_plants	x	x	08_gas	x
            # 18_CT	09_total_transformation_sector	09_06_gas_processing_plants	09_06_02_liquefaction_regasification_plants	x	x	19_total	x
            # 18_CT	09_total_transformation_sector	09_06_gas_processing_plants	x	x	x	08_gas	x
            # 18_CT	09_total_transformation_sector	09_06_gas_processing_plants	x	x	x	19_total	x
            missing_rows_exceptions_dict['gas_processing_ct_total'] = {'_merge':'new_layout_df', 'economy':'18_CT', 'sectors':'09_total_transformation_sector', 'sub1sectors':'09_06_gas_processing_plants', 'fuels':'19_total'}
            
            missing_rows_exceptions_dict['russia_nonspecified_transformation'] = {'_merge':'new_layout_df', 'economy':'16_RUS', 'sectors':'09_total_transformation_sector', 'sub1sectors':'09_12_nonspecified_transformation'}
            
            missing_rows_exceptions_dict['new_statistical_discrepancys'] = {'_merge':'new_layout_df', 'sectors':'11_statistical_discrepancy'}
            
            missing_rows_exceptions_dict['new_22_demand_supply_discrepancy'] = {'_merge':'new_layout_df', 'sectors':'22_demand_supply_discrepancy'}
            
            # 11_geothermal 02_imports#i dont know why these started but they are there :(
            missing_rows_exceptions_dict['geothermal_imports'] = {'_merge':'new_layout_df', 'fuels':'11_geothermal', 'subfuels':'x'}
            # 02_imports	x	x	x	x	12_solar similar to geothermal
            missing_rows_exceptions_dict['imports_solar'] = {'_merge':'new_layout_df', 'sectors':'02_imports', 'fuels':'12_solar'}
            # 03_exports
            missing_rows_exceptions_dict['exports_solar'] = {'_merge':'new_layout_df', 'sectors':'03_exports', 'fuels':'12_solar'}

            #china gas works gas 
            #removeing gas works and biogas to see what happens wehn we do
            
            # missing_rows_exceptions_dict['gas_works_gas_prc'] = {'economy':'05_PRC', '_merge':'new_layout_df', 'subfuels':'08_03_gas_works_gas'}
            # missing_rows_exceptions_dict['biogas_prc'] = {'economy':'05_PRC', '_merge':'new_layout_df', 'subfuels':'16_01_biogas'}
            #created in adjust_projected_supply_to_balance_demand(). for new fuels especially they might turn up and cause issues
            #             09_total_transformation_sector	09_12_nonspecified_transformation	x	x	x	08_gas	08_01_natural_gas
            # 09_total_transformation_sector	09_12_nonspecified_transformation	x	x	x	08_gas	x
            # 09_total_transformation_sector	09_12_nonspecified_transformation	x	x	x	08_gas	08_01_natural_gas
            # 09_total_transformation_sector	09_12_nonspecified_transformation	x	x	x	08_gas	x

            #21_VN	09_total_transformation_sector	09_01_electricity_plants	09_01_06_biomass	x	x	15_solid_biomass	15_01_fuelwood_and_woodwaste
            # 21_VN	09_total_transformation_sector	09_01_electricity_plants	09_01_06_biomass	x	x	15_solid_biomass	15_02_bagasse
            #these two are linked to creating new rows within the layout df to handle newly created, more disaggregated rows in the power reslts.
            # missing_rows_exceptions_dict['VN_09_total_transformation_sector_fuelwood'] = {'economy':'21_VN', 'sectors':'09_total_transformation_sector', 'sub1sectors':'09_01_electricity_plants', 'sub2sectors':'09_01_06_biomass', 'fuels':'15_solid_biomass', 'subfuels':'15_01_fuelwood_and_woodwaste'}
            # missing_rows_exceptions_dict['VN_09_total_transformation_sector_bagasse'] = {'economy':'21_VN', 'sectors':'09_total_transformation_sector', 'sub1sectors':'09_01_electricity_plants', 'sub2sectors':'09_01_06_biomass', 'fuels':'15_solid_biomass', 'subfuels':'15_02_bagasse'}
            ############################
            #use the keys as column names to remove the rows in the dict:
            # for ignored_issue in missing_rows_exceptions_dict.keys():
            #     #iterate through the dict to thin down to the rows we want to remove and then remove them by index
            #     rows_to_remove = missing_rows.copy()
            #     for col in missing_rows_exceptions_dict[ignored_issue].keys():
            #         rows_to_remove = rows_to_remove[rows_to_remove[col] == missing_rows_exceptions_dict[ignored_issue][col]].copy()
            #     missing_rows = missing_rows.drop(rows_to_remove.index).copy()
            
            for ignored_issue, criteria in missing_rows_exceptions_dict.items():
                query = ' & '.join([f"{col} == '{val}'" for col, val in criteria.items()])
                rows_to_remove = missing_rows.query(query).index
                missing_rows.drop(rows_to_remove, inplace=True)

            ###############
            if missing_rows.shape[0] > 0:
                #put them in order to help see the issue
                missing_rows.sort_values(by=shared_categories_old, inplace=True)
                #put the _merge col at front
                missing_rows = missing_rows[['_merge'] + missing_rows.columns[:-1].tolist()]
                economy=missing_rows['economy'].unique()[0]
                missing_rows.to_csv(f'data/temp/error_checking/missing_rows_{economy}.csv', index=False)
                print("There are {} rows where the results file is missing rows from the layout file. These rows have been saved to data/temp/error_checking/missing_rows.csv".format(missing_rows.shape[0]))
                # breakpoint()
        if (merged_df_bad_values.shape[0] > 0 or missing_rows.shape[0] > 0) and not NEW_YEARS_IN_INPUT:
            
            if missing_rows[(missing_rows['subfuels'].str.contains('unallocated'))].shape[0] == missing_rows.shape[0]:
                #if all rows are tehse rows then skip tehse rows since they are part of the solution in allocate_16_15_subfuel_x_rows_to_unallocated() and of course there wont be corresponding rows in the layout df. that is ok!
                pass
            else:
                #save the results_layout_df for user to check
                breakpoint()
                economy = results_layout_df['economy'].unique()[0]
                results_layout_df.to_csv(f'data/temp/error_checking/results_layout_df_{economy}.csv', index=False)
                missing_rows.to_csv(f'data/temp/error_checking/missing_rows_{economy}.csv', index=False)
                raise Exception("The layout df and the newly processed layout df do not match for the years in the layout file. This should not happen.")
        

def process_sheet(sheet_name, excel_file, economy, OUTLOOK_BASE_YEAR, OUTLOOK_LAST_YEAR, mapping_dict):
    wb = load_workbook(filename=excel_file)
    sheet = wb[sheet_name]

    sheet_data = pd.DataFrame()
    missing_entries = []

    for scenario in ['REF', 'TGT']:
        # Initialize variables for the scenario range
        scenario_start_row = None
        scenario_end_row = None

        # Iterate through merged cell ranges to find the scenario range
        for merged_range in sheet.merged_cells.ranges:
            for row in sheet.iter_rows(min_row=merged_range.min_row, max_row=merged_range.max_row, min_col=merged_range.min_col, max_col=merged_range.max_col):
                if row[0].value == scenario:
                    scenario_start_row = merged_range.min_row
                    scenario_end_row = merged_range.max_row
                    break
            if scenario_start_row:
                break

        if not scenario_start_row:
            raise Exception(f"'{scenario}' merged cell range not found in {sheet_name}.")

        # Locate 'Energy Demand (PJ)' cell within the scenario range
        energy_demand_cell = None
        for row in range(scenario_start_row, scenario_end_row + 1):
            for cell in sheet[row]:
                if cell.value == 'Energy Demand (PJ)':
                    energy_demand_cell = cell
                    break
            if energy_demand_cell:
                break

        if not energy_demand_cell:
            raise Exception(f"'Energy Demand (PJ)' cell not found for {scenario} in {sheet_name}.")

        # Calculate the range to read from the Excel file
        start_col = energy_demand_cell.column_letter
        start_row = energy_demand_cell.row

        # Calculate the ending column letter based on the number of years
        end_col_num = energy_demand_cell.column + 81  # 81 additional columns after the start column (1990-2070)
        end_col_letter = get_column_letter(end_col_num)
        
        data_df = read_excel_with_bounds_check(excel_file, sheet_name, start_row, start_col, end_col_letter)
            
        # Drop rows after 'Total'
        energy_demand_header = data_df.columns[0]
        total_index = data_df.index[data_df[energy_demand_header] == 'Total'].tolist()
        if total_index:
            data_df = data_df.iloc[:total_index[0]].reset_index(drop=True)

        # Process the data
        transformed_data = pd.DataFrame()
        for index, row in data_df.iterrows():
            if row[energy_demand_header] == 'Total':
                continue

            mapped_values = mapping_dict.get(row[energy_demand_header], {'fuels': 'Unknown', 'subfuels': 'Unknown'})
            if mapped_values == {'fuels': 'Unknown', 'subfuels': 'Unknown'}:
                missing_entries.append(row[energy_demand_header])

            new_row = {
                'scenarios': 'reference' if scenario == 'REF' else 'target',
                'economy': economy[0],
                'sectors': '16_other_sector',
                'sub1sectors': '16_02_agriculture_and_fishing',
                'sub2sectors': 'x',
                'sub3sectors': 'x',
                'sub4sectors': 'x',
                'fuels': mapped_values['fuels'],
                'subfuels': mapped_values['subfuels'],
                **{str(year): row[year] for year in range(OUTLOOK_BASE_YEAR + 1, OUTLOOK_LAST_YEAR + 1)}
            }
            transformed_data = pd.concat([transformed_data, pd.DataFrame([new_row])], ignore_index=True)
            # transformed_data = transformed_data.append(new_row, ignore_index=True)

        sheet_data = pd.concat([sheet_data, transformed_data])

    # Save missing entries to a DataFrame and CSV
    if missing_entries:
        missing_df = pd.DataFrame(missing_entries, columns=['Missing Entries'])
        missing_df.to_csv('data/temp/error_checking/agriculture_missing_entries.csv', index=False)
        breakpoint()
        raise Exception(f"Missing entries found in {sheet_name}.")

    return sheet_data

def process_agriculture(excel_file, shared_categories, economy, OUTLOOK_BASE_YEAR, OUTLOOK_LAST_YEAR):
    wb = load_workbook(filename=excel_file)

    # Load the mapping document
    mapping_df = pd.read_excel('./config/agriculture_mapping.xlsx')
    mapping_dict = mapping_df.set_index('Energy Demand (PJ)').to_dict('index')

    all_transformed_data = pd.DataFrame()

    # Determine which sheets to process
    if 'Output' in wb.sheetnames:
        all_transformed_data = pd.concat([all_transformed_data, process_sheet('Output', excel_file, economy, OUTLOOK_BASE_YEAR, OUTLOOK_LAST_YEAR, mapping_dict)])
    
    elif 'Agriculture Output' in wb.sheetnames and 'Fishing Output' not in wb.sheetnames:
        all_transformed_data = pd.concat([all_transformed_data, process_sheet('Agriculture Output', excel_file, economy, OUTLOOK_BASE_YEAR, OUTLOOK_LAST_YEAR, mapping_dict)])

    elif 'Agriculture Output' in wb.sheetnames and 'Fishing Output' in wb.sheetnames:
        agri_data = process_sheet('Agriculture Output', excel_file, economy, OUTLOOK_BASE_YEAR, OUTLOOK_LAST_YEAR, mapping_dict)
        fish_data = process_sheet('Fishing Output', excel_file, economy, OUTLOOK_BASE_YEAR, OUTLOOK_LAST_YEAR, mapping_dict)

        # Concatenate the two DataFrames
        combined_data = pd.concat([agri_data, fish_data])

        # Group by shared categories and sum the agri and fish values
        combined_data = combined_data.groupby(shared_categories, as_index=False)[[str(year) for year in range(OUTLOOK_BASE_YEAR + 1, OUTLOOK_LAST_YEAR + 1)]].sum()

        all_transformed_data = pd.concat([all_transformed_data, combined_data])

    # Reorder the columns to match shared_categories
    final_columns = shared_categories + [str(year) for year in range(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR+1)]
    all_transformed_data = all_transformed_data[final_columns]

    # Save the combined transformed data
    # all_transformed_data.to_csv('data/temp/error_checking/agriculture_transformed.csv', index=False)

    return all_transformed_data

def split_fuels_into_subfuels_based_on_historical_splits(csv_file, layout_df, shared_categories, OUTLOOK_BASE_YEAR, OUTLOOK_LAST_YEAR):
    #this will be used to split the fuels into subfuels based on historical splits.
    df = pd.read_csv(csv_file)

    # Determine the fuels to check based on file name
    if 'pipeline' in csv_file.lower():
        fuels_to_check = ['08_gas', '07_petroleum_products']
    # elif 'buildings' in csv_file.lower():
    #     fuels_to_check = ['12_solar']

    # Exclude 'subfuels' from shared_categories
    categories_for_matching = [cat for cat in shared_categories if cat != 'subfuels']

    # Define year columns for analysis (past 5 years)
    year_columns_for_analysis = [str(col) for col in range(OUTLOOK_BASE_YEAR-4, OUTLOOK_BASE_YEAR+1)]

    for fuel in fuels_to_check:
        # Check if the fuel is in the 'fuels' column
        if fuel in df['fuels'].values:
            # Find rows in df with the specific fuel
            fuel_rows = df[df['fuels'] == fuel]

            for _, fuel_row in fuel_rows.iterrows():
                # Extract shared categories values from this row, excluding 'subfuels'
                category_values = {cat: fuel_row[cat] for cat in categories_for_matching}

                # Find rows in layout_df that match these category values exactly
                matching_condition = (layout_df[categories_for_matching] == pd.Series(category_values)).all(axis=1)
                matching_rows = layout_df[matching_condition]

                # Check if the year columns exist in matching_rows
                if not all(col in matching_rows.columns for col in year_columns_for_analysis):
                    raise Exception("Missing year columns in matching_rows:", [col for col in year_columns_for_analysis if col not in matching_rows.columns])

                # Melt matching_rows and keep only the past 5 years
                melted = matching_rows.melt(id_vars=shared_categories, value_vars=year_columns_for_analysis)

                # Sum up the years using groupby
                summed = melted.groupby(shared_categories).sum(numeric_only=True).reset_index()
                # Calculate the sum of those sums:
                sum_of_subfuels = summed.value.sum()
                    
                proportion_dict = {}
                for _, row in summed.iterrows():
                    if row['subfuels'] != 'x':
                        # proportion = row['value'] / total_values.iloc[0]
                        try:
                            proportion = row['value'] / sum_of_subfuels if sum_of_subfuels != 0 else 0
                        except:
                            breakpoint()
                            proportion = row['value'] / sum_of_subfuels if sum_of_subfuels != 0 else 0
                            
                        proportion_dict[row['subfuels']] = proportion

                # Create new rows in df using the proportions
                total_row_df = fuel_rows[fuel_rows['subfuels'] == 'x']
                if not total_row_df.empty:
                    total_row = total_row_df.iloc[0]
                    for subfuel, proportion in proportion_dict.items():
                        new_row = total_row.copy()
                        new_row['subfuels'] = subfuel
                        for year in range(OUTLOOK_BASE_YEAR, OUTLOOK_LAST_YEAR+1):
                            new_row[str(year)] = new_row[str(year)] * proportion
                        # Append the new row to df
                        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)    
                        # df = df.append(new_row, ignore_index=True)
        # Drop the total rows (with 'x' in 'subfuels') for the current fuel type
        df = df.drop(df[(df['fuels'] == fuel) & (df['subfuels'] == 'x')].index)

    return df

import string
def read_excel_with_bounds_check(excel_file, sheet_name, start_row, start_col, end_col_letter):
    # Read the entire sheet to determine the available columns
    full_df = pd.read_excel(excel_file, sheet_name=sheet_name, header=start_row - 1)
    available_columns = full_df.columns

    try:
        # Convert end_col_letter to column index, if it is a letter
        if isinstance(end_col_letter, str):
            end_col_index = available_columns.get_loc(end_col_letter)
        else:
            end_col_index = len(available_columns) - 1
    except KeyError:
        print(f"Warning: In Agriculture data, end column '{end_col_letter}' is not available. Defaulting to last column.")
        end_col_index = len(available_columns) - 1

    # Ensure column indices are within bounds
    if end_col_index >= len(available_columns):
        end_col_index = len(available_columns) - 1

    # Convert column indices back to Excel letter range for usecols
    if end_col_index // 26 == 0:
        end_col_letter = string.ascii_uppercase[end_col_index]
    else:
        end_col_letter = string.ascii_uppercase[end_col_index // 26 - 1] + string.ascii_uppercase[end_col_index % 26]
    column_range = f"{start_col}:{end_col_letter}"
        
    # Read the data from the Excel file with the adjusted usecols parameter
    data_df = pd.read_excel(excel_file, sheet_name=sheet_name, header=start_row - 1, usecols=column_range)

    return data_df


def insert_data_centres_into_layout_df(layout_df, results_df, shared_categories):
    # #teams:
    # Testing removing it from the <=2022 data within merging script. seems like its working fine. 
    # Leanne i think the best way is that you will need to still do the calculation on your side so the data you give me takes in to account the effect of data centres and ai training on services in all years. 
    # I think also a good improvement would be to have it so that when you do the calcaltion you also merge the data centres and ai training data into the service sectors data and give me the results in one file. That way sicne the integration script takes in each modellers file iteratively and checks against the data it has for pre-2022, it can check that the reaon why buildings pre-2022 data is different to wat it has is because of the data centres data (which it would be a little more complicated to do if we provided the data centres data in a separate file).        
    # And then ultimate goal imo would be to have you run data centres model yourself so that there is no passing between modellers which i dont like because it slows thing down and creates extra steps in peoples heads.
    
    #before we remove all non results years, if  16_01_03_ai_training or 16_01_04_traditional_data_centres are in the sub2sectors column then take their values away from 16_01_01_commercial_and_public_services (where fuel is 17_electricity) in the layout data and add them in. 
    if '16_01_03_ai_training' in results_df['sub2sectors'].unique() or '16_01_04_traditional_data_centres' in results_df['sub2sectors'].unique():
        #so for layout_df, add these row, but only for years <= OUTLOOK_BASE_YEAR (for the otehrs, set value to 0). THen minus these values from 16_01_01_commercial_and_public_services.
        layout_df_service_sectors = layout_df.loc[(layout_df['sub2sectors'] == '16_01_01_commercial_and_public_services') &(layout_df['fuels'] == '17_electricity')].copy()
        #drop tehse rows from the layout_df as well as the rows for data centres
        layout_df = layout_df.loc[~((layout_df['sub2sectors'].isin(['16_01_03_ai_training', '16_01_04_traditional_data_centres','16_01_01_commercial_and_public_services'])) &(layout_df['fuels'] == '17_electricity'))].copy()
        
        #melt the years
        layout_df_service_sectors = layout_df_service_sectors.melt(id_vars=shared_categories+['is_subtotal'], var_name='year', value_name='value')
        #sum up the data centres values and then join them
        data_centres_df = results_df.loc[(results_df['sub2sectors'].isin(['16_01_03_ai_training', '16_01_04_traditional_data_centres'])  & (results_df['fuels'] == '17_electricity'))].copy()
        #and if & (results_df['subtotal_layout'] == False) & (results_df['subtotal_results'] == False) are in the df, then filter for the too:
        if 'subtotal_layout' in data_centres_df.columns:
            data_centres_df = data_centres_df.loc[(data_centres_df['subtotal_layout'] == False)].copy()
            #drop subtotals cols
            data_centres_df.drop(columns=['subtotal_layout'], inplace=True)
        if 'subtotal_results' in data_centres_df.columns:
            data_centres_df = data_centres_df.loc[(data_centres_df['subtotal_results'] == False)].copy()
            #drop subtotals cols
            data_centres_df.drop(columns=['subtotal_results'], inplace=True)
        data_centres_df_melt = data_centres_df.melt(id_vars=shared_categories, var_name='year', value_name='value')
        #make year as int
        data_centres_df_melt['year'] = data_centres_df_melt['year'].astype(int)
        layout_df_service_sectors['year'] = layout_df_service_sectors['year'].astype(int)
        #set non layout years to 0. this will also be what we add as new layout rows
        data_centres_df_melt.loc[data_centres_df_melt.year > OUTLOOK_BASE_YEAR, 'value'] = 0
        #sum up the data centres values so we can subtract them from the service sectors values
        data_centres_df_sum = data_centres_df_melt.groupby(['scenarios', 'year']).sum(numeric_only=True).reset_index()
        
        layout_df_service_sectors = layout_df_service_sectors.merge(data_centres_df_sum, on=['scenarios', 'year'], suffixes=('', '_results'), how='left')
        #now minus the data centres values from the service sectors values
        layout_df_service_sectors['value'] = layout_df_service_sectors['value'] - layout_df_service_sectors['value_results'].fillna(0)
        #if any negative vlaues raise an error
        if (layout_df_service_sectors['value'] < 0).any():
            breakpoint()
            raise ValueError('Negative values found when subtracting data centres from service sectors.')
        #now drop the results cols and add the new rows to the layout_df after cleaning up
        layout_df_service_sectors.drop(columns=[col for col in layout_df_service_sectors.columns if '_results' in col], inplace=True)
        #pivot
        layout_df_service_sectors = layout_df_service_sectors.pivot(index=shared_categories+['is_subtotal'], columns='year', values='value').reset_index()
        #make sure all years are strs
        layout_df_service_sectors.columns = [str(col) for col in layout_df_service_sectors.columns]
        #and clean up the data_centres_df_melt by pivoting and then setting subtotal to False
        data_centres_df_layout = data_centres_df_melt.pivot(index=shared_categories, columns='year', values='value').reset_index()
        
        #make sure all years are strs
        data_centres_df_layout.columns = [str(col) for col in data_centres_df_layout.columns]
        data_centres_df_layout['is_subtotal'] = False
        #concatenate all the enw layout rows
        layout_df = pd.concat([layout_df, layout_df_service_sectors, data_centres_df_layout])
        #now carry on (:
    return layout_df

def save_merged_file(final_df, SINGLE_ECONOMY_ID, previous_merged_df_filename, shared_categories_w_subtotals, 
    folder_path, old_folder_path,   
    COMPARE_TO_PREVIOUS_MERGE = True):
    """Saves the merged file to a new CSV file and moves the previous merged file to the 'old' folder."""
    # Identify the previous merged file
    
    # Check if the folder already exists
    if not os.path.exists(folder_path) and (isinstance(SINGLE_ECONOMY_ID, str)):
        # If the folder doesn't exist, create it
        os.makedirs(folder_path)
    
    # Check if the old folder exists
    if not os.path.exists(old_folder_path):
        # If the old folder doesn't exist, create it
        os.makedirs(old_folder_path)
    
    if COMPARE_TO_PREVIOUS_MERGE:
        compare_to_previous_merge(final_df, shared_categories_w_subtotals, results_data_path=folder_path,previous_merged_df_filename=previous_merged_df_filename, new_subtotal_columns=['subtotal_layout', 'subtotal_results'], previous_subtotal_columns=['subtotal_historic','subtotal_predicted','subtotal'])
    
    previous_merged_df_filename = None
    if os.path.exists(folder_path):
        for file in os.listdir(folder_path):
            if file.startswith(f'merged_file_energy_{SINGLE_ECONOMY_ID}') and file.endswith('.csv'):
                previous_merged_df_filename = file
                break

    # Move the old merged file to the 'old' folder if it exists
    if previous_merged_df_filename:
        old_file_path = f'{folder_path}/{previous_merged_df_filename}'
        new_old_file_path = f'{old_folder_path}/{previous_merged_df_filename}'
        
        # Remove the old file in the 'old' folder if it exists
        if os.path.exists(new_old_file_path):
            os.remove(new_old_file_path)
        
        os.rename(old_file_path, new_old_file_path)

    #save the combined data to a new Excel file
    #layout_df.to_excel('../../tfc/combined_data.xlsx', index=False, engine='openpyxl')
    date_today = datetime.now().strftime('%Y%m%d')
    if (isinstance(SINGLE_ECONOMY_ID, str)):
        final_df.to_csv(f'{folder_path}/merged_file_energy_{SINGLE_ECONOMY_ID}_{date_today}.csv', index=False)
    else:
        final_df.to_csv(f'results/merged_file_energy_{date_today}.csv', index=False)